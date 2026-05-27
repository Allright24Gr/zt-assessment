"""manual-checklist.xlsx 재생성 — 담당자 선택 컬럼을 O/△/X/평가불가 dropdown 으로 통일.

변경 사항:
  · manual_diagnosis: G열(★ 담당자 선택)에 DataValidation 추가
                      (O/△/X/평가불가). 헤더 라벨/안내문 업데이트.
  · judgment_mapping: 선택값 → 판정 1:1 매핑 (참고용).
  · instructions:     신규 평가 규칙 안내로 교체.

기존 파일을 in-place 로 덮어쓰고 .bak 백업을 남긴다.

실행: python backend/scripts/regenerate_manual_checklist.py
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "manual-checklist.xlsx"
BAK  = ROOT / "manual-checklist.xlsx.bak"

CHOICE_OPTIONS = ["O", "△", "X", "평가불가"]
CHOICE_TO_VERDICT = {
    "O":      "충족",
    "△":      "부분충족",
    "X":      "미충족",
    "평가불가": "평가불가",
}


def main() -> int:
    if not XLSX.exists():
        print(f"[err] 파일 없음: {XLSX}", file=sys.stderr)
        return 1

    if not BAK.exists():
        shutil.copy2(XLSX, BAK)
        print(f"[ok] 백업 생성: {BAK.name}")
    else:
        print(f"[skip] 백업 이미 존재: {BAK.name}")

    wb = openpyxl.load_workbook(XLSX)

    # ── 1) manual_diagnosis ──────────────────────────────────────────────
    ws_diag = wb["manual_diagnosis"]

    ws_diag.cell(2, 1).value = (
        "★ 담당자 선택(G열)에서 각 항목 질문에 대한 평가 결과를 선택하세요. "
        "O = 충족(질문 부합) / △ = 부분충족 / X = 미충족 / 평가불가 = 측정 불가. "
        "양식은 모든 항목 동일 선택지로 통일되어 있습니다."
    )
    ws_diag.cell(3, 7).value = "★ 담당자 선택 (필수)"

    # 기존 DataValidations 제거 후 새로 적용
    ws_diag.data_validations.dataValidation = []

    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(CHOICE_OPTIONS)}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="선택값 확인",
        error="O / △ / X / 평가불가 중에서 선택해주세요.",
    )

    data_row_count = 0
    for r in range(4, ws_diag.max_row + 1):
        mid = ws_diag.cell(r, 1).value
        if not mid or str(mid).startswith("▸"):
            continue
        if not str(mid).startswith("M"):
            continue
        dv.add(f"G{r}")
        # 기존 자유 텍스트 입력값 비움 (드롭다운 위배 방지)
        existing = ws_diag.cell(r, 7).value
        if existing and str(existing).strip() not in CHOICE_OPTIONS:
            ws_diag.cell(r, 7).value = None
        data_row_count += 1
    ws_diag.add_data_validation(dv)
    print(f"[ok] manual_diagnosis: {data_row_count} 행에 dropdown 적용")

    # ── 2) judgment_mapping 재구성 (참고용 1:1 매핑) ──────────────────────
    if "judgment_mapping" in wb.sheetnames:
        del wb["judgment_mapping"]
    ws_judg = wb.create_sheet("judgment_mapping")

    ws_judg.cell(1, 1).value = (
        "판정 매핑 — 담당자 선택 → 판정 결과 (참고용, 모든 항목 공통)"
    )
    ws_judg.cell(1, 1).font = Font(name="Arial", size=11, bold=True, color="FF1E3A5F")
    ws_judg.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    header_row = 2
    headers = ["담당자 선택", "판정 결과"]
    for c, h in enumerate(headers, start=1):
        cell = ws_judg.cell(header_row, c)
        cell.value = h
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF1E3A5F")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, choice in enumerate(CHOICE_OPTIONS):
        ws_judg.cell(header_row + 1 + i, 1).value = choice
        ws_judg.cell(header_row + 1 + i, 2).value = CHOICE_TO_VERDICT[choice]
        ws_judg.cell(header_row + 1 + i, 1).alignment = Alignment(horizontal="center")
        ws_judg.cell(header_row + 1 + i, 2).alignment = Alignment(horizontal="center")

    ws_judg.column_dimensions["A"].width = 16
    ws_judg.column_dimensions["B"].width = 16

    print(f"[ok] judgment_mapping: {len(CHOICE_OPTIONS)} 행 (1:1 매핑)")

    # ── 3) instructions 교체 ──────────────────────────────────────────────
    if "instructions" in wb.sheetnames:
        del wb["instructions"]
    ws_inst = wb.create_sheet("instructions")

    rows = [
        ("작성 안내 — Instructions",                                                                                ""),
        ("",                                                                                                         ""),
        ("1단계", "manual_diagnosis 시트의 '★ 담당자 선택 (필수)' 열에서 각 항목의 평가 결과를 선택하세요."),
        ("",      "선택지: O / △ / X / 평가불가 (드롭다운)"),
        ("2단계", "비고/증적메모 열에 증빙 자료 설명이나 파일명을 기록하세요(선택사항)."),
        ("3단계", "작성 완료 후 웹 대시보드에서 이 파일을 업로드하면 자동으로 DB에 저장됩니다."),
        ("4단계", "판정 결과는 다음 매핑으로 자동 산출됩니다:"),
        ("",      "  · O      → 충족 (해당 항목 평가 기준에 부합)"),
        ("",      "  · △      → 부분충족 (일부 부합 / 운영은 되지만 범위·증거 제한)"),
        ("",      "  · X      → 미충족 (해당 통제 없음 또는 운영 안 됨)"),
        ("",      "  · 평가불가 → 평가불가 (외부에서 측정 불가 / 권한 부족 등)"),
        ("",      ""),
        ("참고",  "항목별 maturity(성숙도)는 D열에 표시되며 점수 가중치에만 영향을 줍니다. 담당자 선택은 모든 항목에서 동일한 4가지 기호 중 하나입니다."),
        ("",      ""),
        ("주의",  "G열 외 다른 셀은 수정하지 마세요. M_id·항목번호·성숙도는 업로드 매칭에 사용됩니다."),
    ]

    for r, (k, v) in enumerate(rows, start=1):
        ws_inst.cell(r, 1).value = k
        ws_inst.cell(r, 2).value = v
        if r == 1:
            ws_inst.cell(r, 1).font = Font(name="Arial", size=12, bold=True, color="FF1E3A5F")
        elif k in ("주의", "참고"):
            ws_inst.cell(r, 1).font = Font(name="Arial", size=10, bold=True, color="FFB91C1C")
        elif k:
            ws_inst.cell(r, 1).font = Font(name="Arial", size=10, bold=True)
        ws_inst.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="center")

    ws_inst.column_dimensions["A"].width = 14
    ws_inst.column_dimensions["B"].width = 110

    print(f"[ok] instructions: {len(rows)} 행")

    # ── 저장 ─────────────────────────────────────────────────────────────
    wb.save(XLSX)
    print(f"[ok] 저장 완료: {XLSX.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
