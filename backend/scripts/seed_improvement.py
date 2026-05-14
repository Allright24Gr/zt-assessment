"""zt-improvement-guide.xlsx → ImprovementGuide 테이블 seed 스크립트.

실행:
    python3 backend/scripts/seed_improvement.py
    (컨테이너 내부) python3 /app/scripts/seed_improvement.py
"""

import sys
import os
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from database import SessionLocal
from models import Checklist, ImprovementGuide

import openpyxl


MATURITY_NUM = {"기존": 1, "초기": 2, "향상": 3, "최적화": 4}

TOOL_PRIORITY = {
    "keycloak": "High",
    "wazuh": "High",
    "nmap": "Medium",
    "trivy": "Medium",
    "수동": "Low",
}

MATURITY_TERM = {
    "기존": "단기",
    "초기": "단기",
    "향상": "중기",
    "최적화": "장기",
}


def _find_xlsx() -> Path:
    candidates = [
        Path("/app/zt-improvement-guide.xlsx"),
        Path(__file__).parent.parent / "zt-improvement-guide.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError("zt-improvement-guide.xlsx 파일을 찾을 수 없습니다.")


def seed():
    xlsx = _find_xlsx()
    print(f"[seed_improvement] {xlsx} 파일 로드 중...")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["improvement_guide"]

    db = SessionLocal()
    inserted = updated = skipped = 0

    try:
        # 체크리스트 item_id → check_id 캐시
        checklist_cache: dict[str, int] = {
            row.item_id: row.check_id
            for row in db.query(Checklist.item_id, Checklist.check_id).all()
        }

        for r in range(3, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, 13)]
            item_id = row[0]

            if not item_id or item_id == "항목ID":
                continue

            has_guide = str(row[11]).strip() if row[11] else ""
            if has_guide != "있음":
                skipped += 1
                continue

            pillar      = str(row[1]).strip() if row[1] else ""
            maturity    = str(row[3]).strip() if row[3] else ""
            tool_raw    = str(row[6]).strip().lower() if row[6] else "수동"
            insuf_text  = str(row[8]).strip() if row[8] else ""
            partial_text = str(row[9]).strip() if row[9] else ""
            solution    = str(row[10]).strip() if row[10] else ""

            check_id = checklist_cache.get(item_id)
            priority = TOOL_PRIORITY.get(tool_raw, "Medium")
            term     = MATURITY_TERM.get(maturity, "중기")

            steps_list = [s.strip() for s in solution.split("\n") if s.strip()] if solution else []

            # 미충족 → 별도 가이드 행
            if insuf_text:
                existing = db.query(ImprovementGuide).filter(
                    ImprovementGuide.check_id == check_id,
                    ImprovementGuide.current_level == maturity,
                    ImprovementGuide.task == insuf_text,
                ).first()
                if existing:
                    existing.expected_gain = partial_text or None
                    existing.steps = steps_list or None
                    existing.recommended_tool = tool_raw
                    updated += 1
                else:
                    db.add(ImprovementGuide(
                        check_id=check_id,
                        pillar=pillar,
                        task=insuf_text,
                        priority=priority,
                        term=term,
                        recommended_tool=tool_raw,
                        current_level=maturity,
                        expected_gain=partial_text or None,
                        steps=steps_list or None,
                        expected_effect=partial_text or None,
                    ))
                    inserted += 1

        db.commit()
        print(f"[seed_improvement] 완료: 삽입={inserted} 업데이트={updated} 건너뜀(권고없음)={skipped}")
    except Exception as e:
        db.rollback()
        print(f"[seed_improvement] 오류: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    seed()
