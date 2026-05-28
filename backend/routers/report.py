import asyncio
import io
import logging
import math
import os
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session

from database import get_db
from models import (
    DiagnosisSession, DiagnosisResult, MaturityScore,
    Checklist, Organization, User, ImprovementGuide, Evidence, CollectedData,
    ScoreHistory,
)
from scoring.engine import determine_maturity_level
from routers.auth import get_current_user, assert_session_access
from routers.assessment import build_evaluation_meta
from services.standards_mapping import map_item_to_standards, session_standards_summary

router = APIRouter()
logger = logging.getLogger("zt.report")

# ── 한글 폰트 등록 (NanumGothic, Dockerfile에서 fonts-nanum 설치) ──────────
_FONT_REGISTERED = False
_FONT_NAME = "Helvetica"  # fallback

def _ensure_font():
    global _FONT_REGISTERED, _FONT_NAME
    if _FONT_REGISTERED:
        return
    candidates = [
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/nanum/NanumGothic.ttf",
    ]
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        for path in candidates:
            if os.path.exists(path):
                pdfmetrics.registerFont(TTFont("NanumGothic", path))
                _FONT_NAME = "NanumGothic"
                break
        else:
            logger.warning(
                "[report] NanumGothic 폰트를 찾지 못해 Helvetica 폴백 사용. 후보: %s",
                candidates,
            )
    except Exception as e:
        logger.warning("[report] 폰트 등록 실패 — Helvetica 폴백 (%s)", e)
    _FONT_REGISTERED = True


def _first_step(steps) -> str:
    """ImprovementGuide.steps 가 list/dict/str 어떤 형태든 첫 step 문자열 반환."""
    if not steps:
        return ""
    if isinstance(steps, list):
        first = steps[0] if steps else ""
        if isinstance(first, dict):
            return str(first.get("description") or first.get("step") or first.get("text") or "")
        return str(first) if first else ""
    if isinstance(steps, dict):
        try:
            keys = sorted(steps.keys())
            return str(steps[keys[0]]) if keys else ""
        except Exception:
            return ""
    return str(steps)


# ── 공통 데이터 빌더 ────────────────────────────────────────────────────────

def _build_data(session_id: int, db: Session) -> dict:
    session = db.query(DiagnosisSession).filter(
        DiagnosisSession.session_id == session_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")

    org  = db.query(Organization).filter(Organization.org_id == session.org_id).first()
    user = db.query(User).filter(User.user_id == session.user_id).first()

    maturity_rows = db.query(MaturityScore).filter(
        MaturityScore.session_id == session_id
    ).all()
    pillar_scores = [
        {
            "pillar": m.pillar,
            "score": round(m.score or 0.0, 3),
            "level": determine_maturity_level(m.score or 0.0),
            "pass_cnt": m.pass_cnt or 0,
            "fail_cnt": m.fail_cnt or 0,
            "na_cnt": m.na_cnt or 0,
        }
        for m in maturity_rows
    ]

    results = (
        db.query(DiagnosisResult, Checklist)
        .join(Checklist, DiagnosisResult.check_id == Checklist.check_id)
        .filter(DiagnosisResult.session_id == session_id)
        .all()
    )

    # 증적/관찰 (CollectedData.raw_json + metric) - 항목별 미리 캐시.
    collected_by_check: dict[int, dict] = {}
    for row in db.query(CollectedData).filter(CollectedData.session_id == session_id).all():
        collected_by_check[row.check_id] = {
            "tool":         row.tool,
            "metric_key":   row.metric_key,
            "metric_value": row.metric_value,
            "threshold":    row.threshold,
            "raw_json":     row.raw_json if isinstance(row.raw_json, dict) else {},
            "error":        row.error,
            "collected_at": row.collected_at.isoformat() if row.collected_at else None,
        }
    # Evidence (수동 입력 관찰) — 항목별 1건씩만 (최신).
    evidence_by_check: dict[int, dict] = {}
    for ev in db.query(Evidence).filter(Evidence.session_id == session_id).all():
        evidence_by_check[ev.check_id] = {
            "source":   ev.source or "",
            "observed": ev.observed or "",
            "location": ev.location or "",
            "reason":   ev.reason or "",
        }

    checklist_results, fail_items = [], []
    for dr, cl in results:
        coll = collected_by_check.get(cl.check_id) or {}
        raw = coll.get("raw_json") or {}
        ev = evidence_by_check.get(cl.check_id) or {}

        evidence_summary_parts: list[str] = []
        if coll.get("metric_key"):
            mv = coll.get("metric_value")
            th = coll.get("threshold")
            evidence_summary_parts.append(
                f"{coll['metric_key']} = {mv}" + (f" / 임계값 {th}" if th is not None else "")
            )
        if raw.get("issues"):
            evidence_summary_parts.append(f"발견 이슈 {len(raw['issues'])}건")
        if ev.get("observed"):
            evidence_summary_parts.append(f"관찰: {ev['observed'][:200]}")
        evidence_summary = " · ".join(evidence_summary_parts) if evidence_summary_parts else ""

        reason_label = raw.get("reason_label") or raw.get("reason_code") or ""
        issues = (raw.get("issues") or [])[:3] if isinstance(raw.get("issues"), list) else []

        item = {
            "item_id":       cl.item_id,
            "check_id":      cl.check_id,
            "pillar":        cl.pillar,
            "category":      cl.category,
            "item_name":     cl.item_name,
            "question":      cl.question or "",
            "maturity":      cl.maturity,
            "maturity_score": cl.maturity_score,
            "diagnosis_type": cl.diagnosis_type,
            "tool":          cl.tool,
            "result":        dr.result,
            "score":         dr.score or 0.0,
            "criteria":      cl.criteria or "",
            "recommendation": dr.recommendation or "",
            "evidence_summary": evidence_summary,
            "evidence_source": ev.get("source", ""),
            "evidence_observed": ev.get("observed", ""),
            "evidence_location": ev.get("location", ""),
            "evidence_reason": ev.get("reason", ""),
            "auto_issues": issues,
            "auto_error": coll.get("error", ""),
            "reason_label": reason_label,
            "collected_at": coll.get("collected_at"),
            # 체크리스트 세부(10장)에서 raw_json 요약 출력
            "raw_json_summary": _summarize_raw_json(raw),
            "metric_key":     coll.get("metric_key", ""),
            "metric_value":   coll.get("metric_value"),
            "threshold":      coll.get("threshold"),
        }
        checklist_results.append(item)
        if dr.result in ("미충족", "부분충족"):
            fail_items.append(item)

    fail_check_ids = [
        dr.check_id for dr, _ in results
        if dr.result in ("미충족", "부분충족")
    ]
    guide_rows = db.query(ImprovementGuide).filter(
        ImprovementGuide.check_id.in_(fail_check_ids)
    ).order_by(ImprovementGuide.priority, ImprovementGuide.term).all() if fail_check_ids else []

    check_meta = {cl.check_id: cl for _, cl in results}

    def _steps_list(steps) -> list[str]:
        if not steps:
            return []
        if isinstance(steps, list):
            return [str(s.get("description") or s.get("step") or s.get("text") or s) if isinstance(s, dict) else str(s) for s in steps]
        if isinstance(steps, dict):
            return [str(v) for _, v in sorted(steps.items())]
        if isinstance(steps, str):
            return [line.strip() for line in steps.split("\n") if line.strip()]
        return [str(steps)]

    improvements = []
    for g in guide_rows:
        cl_meta = check_meta.get(g.check_id)
        improvements.append({
            "pillar":         g.pillar,
            "category":       cl_meta.category if cl_meta else "",
            "item_id":        cl_meta.item_id if cl_meta else "",
            "task":           g.task,
            "priority":       g.priority,
            "term":           g.term,
            "tool":           g.recommended_tool or "",
            "current_level":  g.current_level or "",
            "expected_gain":  g.expected_gain or "",
            "expected_effect": g.expected_effect or "",
            "steps":          _steps_list(g.steps),
            "solution":       _first_step(g.steps),
        })

    total = len(checklist_results)
    pass_cnt    = sum(1 for r in checklist_results if r["result"] == "충족")
    partial_cnt = sum(1 for r in checklist_results if r["result"] == "부분충족")
    fail_cnt    = sum(1 for r in checklist_results if r["result"] == "미충족")
    na_cnt      = sum(1 for r in checklist_results if r["result"] == "평가불가")
    overall_score = session.total_score or 0.0
    overall_level = session.level or determine_maturity_level(overall_score)

    # 추이: 같은 org의 ScoreHistory 시간순
    history_rows = (
        db.query(ScoreHistory)
        .filter(ScoreHistory.org_id == session.org_id)
        .order_by(ScoreHistory.assessed_at.asc())
        .all()
    )
    history = [
        {
            "session_id":     h.session_id,
            "assessed_at":    h.assessed_at.isoformat() if h.assessed_at else "",
            "total_score":    round(h.total_score or 0.0, 3),
            "maturity_level": h.maturity_level,
            "pillar_scores":  h.pillar_scores if isinstance(h.pillar_scores, dict) else {},
        }
        for h in history_rows
    ]

    # 직전 세션 (현재 세션 제외)
    prev_hist = None
    for h in history:
        if h["session_id"] != session_id:
            if prev_hist is None or h["assessed_at"] > prev_hist["assessed_at"]:
                prev_hist = h

    # 표준 매핑 요약
    cl_for_std = [
        {
            "item_id":  r["item_id"],
            "pillar":   r["pillar"],
            "category": r["category"],
            "item":     r["item_name"],
            "result":   r["result"],
        }
        for r in checklist_results
    ]
    std_summary = session_standards_summary(cl_for_std) if checklist_results else {
        "nist_800_207": [], "cis_controls_v8": [],
    }

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "session": {
            "session_id":   session.session_id,
            "org":          org.name if org else "",
            "manager":      user.name if user else "",
            "started_at":   session.started_at.isoformat() if session.started_at else "",
            "completed_at": session.completed_at.isoformat() if session.completed_at else "",
            "status":       session.status,
        },
        "summary": {
            "overall_score": round(overall_score, 3),
            "overall_level": overall_level,
            "total_items":   total,
            "pass_cnt":      pass_cnt,
            "partial_cnt":   partial_cnt,
            "fail_cnt":      fail_cnt,
            "na_cnt":        na_cnt,
            "pass_rate":     round(pass_cnt / total, 3) if total > 0 else 0.0,
            "confidence":    round((total - na_cnt) / total, 4) if total > 0 else 0.0,
        },
        "pillar_scores":      pillar_scores,
        "checklist_results":  checklist_results,
        "improvement_targets": fail_items,
        "improvements":       improvements,
        "history":            history,
        "previous_session":   prev_hist,
        "standards_summary":  std_summary,
        "evaluation_meta":    build_evaluation_meta(session),
    }


def _summarize_raw_json(raw: dict, max_len: int = 280) -> str:
    """raw_json 을 보고서용 짧은 dict-like 문자열로 요약."""
    if not isinstance(raw, dict) or not raw:
        return ""
    keep_keys = (
        "score", "metric_value", "threshold", "count", "total", "ratio",
        "issues", "passed", "failed", "enabled", "registered",
        "reason_label", "reason_code", "version", "status",
    )
    pieces = []
    for k in keep_keys:
        if k in raw:
            v = raw[k]
            if isinstance(v, list):
                pieces.append(f'"{k}": {len(v)}건')
            elif isinstance(v, dict):
                pieces.append(f'"{k}": {{…}}')
            else:
                pieces.append(f'"{k}": {v}')
    if not pieces:
        # 첫 3개 key만 노출
        for i, (k, v) in enumerate(raw.items()):
            if i >= 3:
                break
            if isinstance(v, (list, dict)):
                pieces.append(f'"{k}": {type(v).__name__}')
            else:
                pieces.append(f'"{k}": {v}')
    text = "{ " + ", ".join(pieces) + " }"
    return text[:max_len]


# ─────────────────────────────────────────────────────────────────────────────
# PDF 생성 — 11페이지 결재용 보고서 (PDF 샘플 그대로)
# ─────────────────────────────────────────────────────────────────────────────

# 색상 토큰
_COLORS = {
    # 등급
    "기존":   "#dc2626",
    "초기":   "#eab308",
    "향상":   "#2563eb",
    "최적화": "#16a34a",
    # 판정
    "충족":    "#16a34a",
    "부분충족": "#d97706",
    "미충족":  "#dc2626",
    "평가불가": "#9ca3af",
    # base
    "ink":       "#111827",
    "muted":     "#6b7280",
    "subtle":    "#9ca3af",
    "border":    "#e5e7eb",
    "border2":   "#cbd5e1",
    "panel":     "#f8fafc",
    "panel2":    "#f1f5f9",
    "navy":      "#1e3a5f",
    "navy_dim":  "#334155",
    "accent":    "#2563eb",
    "accent_lt": "#dbeafe",
}


def _ko_pillar_short(pillar: str) -> str:
    """레이더 차트용 짧은 라벨."""
    m = {
        "식별자 및 신원":         "식별자",
        "기기 및 엔드포인트":     "기기",
        "네트워크":               "네트워크",
        "시스템":                 "시스템",
        "애플리케이션 및 워크로드": "앱",
        "데이터":                 "데이터",
    }
    return m.get(pillar, pillar[:5])


def _doc_no(session_id: int, generated_at_iso: str) -> str:
    """문서 번호 RZT-YYYY-MMDD-NNN."""
    try:
        d = datetime.fromisoformat(generated_at_iso.replace("Z", "+00:00"))
    except Exception:
        d = datetime.now(timezone.utc)
    return f"RZT-{d.strftime('%Y-%m%d')}-{session_id:03d}"


def _level_for(score: float) -> str:
    return determine_maturity_level(score or 0.0)


def _part_badge_flowable(text: str, font: str):
    """부 마커 — 작은 파란 둥근 배지 안에 흰 글자.
    샘플 PDF의 '1부 · 문서 식별' 스타일.
    """
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.lib import colors
    from reportlab.pdfbase.pdfmetrics import stringWidth

    PADDING_X = 10
    PADDING_Y = 5
    SIZE = 10
    text_w = stringWidth(text, font, SIZE)
    W = text_w + PADDING_X * 2
    H = SIZE + PADDING_Y * 2 + 2

    d = Drawing(W, H)
    d.add(Rect(0, 0, W, H, fillColor=colors.HexColor("#dbeafe"),
               strokeColor=colors.HexColor("#93c5fd"), strokeWidth=0.5, rx=4, ry=4))
    d.add(String(W / 2, PADDING_Y + 2, text, textAnchor="middle",
                 fontName=font, fontSize=SIZE, fillColor=colors.HexColor("#1e3a5f")))
    return d


def _section_heading_flowable(number: str, title: str, font: str, total_w: float):
    """장 제목 — 좌측 굵은 파란 막대 + 검정 큰 제목.
    예: '4. 진단 범위 및 방법론' (좌측 4pt 두께 파란 막대 + 옆 큰 제목)
    """
    from reportlab.platypus import Table, Paragraph
    from reportlab.platypus.tables import TableStyle
    from reportlab.lib import colors

    style_title = ParagraphStyle(
        "_sec_title", fontName=font, fontSize=15, leading=20,
        textColor=colors.HexColor("#0f172a"), leftIndent=0,
    )
    label = Paragraph(f"<b>{number}. {title}</b>", style_title)
    # 좌측 4pt 파란 막대를 Table cell의 BACKGROUND로
    t = Table([["", label]], colWidths=[5, total_w - 5],
              rowHeights=[24])
    t.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (0, 0), colors.HexColor("#1e3a5f")),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (1, 0), (1, 0), 10),
        ("RIGHTPADDING", (1, 0), (1, 0), 0),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
    ]))
    return t


_LEVEL_BADGE_COLORS = {
    "기존":   ("#fee2e2", "#991b1b"),   # 빨강
    "초기":   ("#fef3c7", "#92400e"),   # 노랑/오렌지
    "향상":   ("#dbeafe", "#1e40af"),   # 파랑
    "최적화": ("#d1fae5", "#065f46"),   # 초록
}

_VERDICT_CHIP_COLORS = {
    "충족":     ("#d1fae5", "#065f46"),
    "부분충족": ("#fef3c7", "#92400e"),
    "미충족":   ("#fee2e2", "#991b1b"),
    "평가불가": ("#e5e7eb", "#475569"),
}


def _cover_stat_card(W_pt: float, H_pt: float, title: str, big_text: str,
                     big_color: str, badge_text: str | None, badge_palette: tuple[str, str] | None,
                     sub_line1: str, sub_line2: str, font: str,
                     big_suffix: str = ""):
    """표지용 통계 카드 (Drawing) — 흰 배경 + 옅은 회색 외곽선.

    레이아웃 (위→아래):
      · title          (작은 회색, y=H-16)
      · big_text       (큰, 가운데, y≈H/2+10)
      · big_suffix     (옆에 작게, 옵션 — "/ 4.0" 처럼)
      · badge          (선택, 그 아래, 노란/파란 등 등급 배지)
      · sub_line1      (작은 회색, 하단)
      · sub_line2      (더 작은 회색, 가장 아래)
    """
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.lib import colors
    from reportlab.pdfbase.pdfmetrics import stringWidth

    d = Drawing(W_pt, H_pt)
    # 외곽 카드 (둥근 모서리 + 옅은 회색 테두리)
    d.add(Rect(0.5, 0.5, W_pt - 1, H_pt - 1, fillColor=colors.white,
               strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.8, rx=6, ry=6))

    # title — 카드 상단
    d.add(String(W_pt / 2, H_pt - 16, title, textAnchor="middle",
                 fontName=font, fontSize=9, fillColor=colors.HexColor("#64748b")))

    # big_text + 옵셔널 suffix (예: 2.43 / 4.0)
    BIG_SIZE = 26
    SUF_SIZE = 11
    big_w = stringWidth(big_text, font, BIG_SIZE)
    suf_w = stringWidth(big_suffix, font, SUF_SIZE) if big_suffix else 0
    gap   = 4 if big_suffix else 0
    total = big_w + gap + suf_w
    start_x = (W_pt - total) / 2

    # baseline 위치 — 카드 중앙보다 살짝 위
    big_y = H_pt / 2 + 6
    d.add(String(start_x, big_y, big_text, textAnchor="start",
                 fontName=font, fontSize=BIG_SIZE, fillColor=colors.HexColor(big_color)))
    if big_suffix:
        d.add(String(start_x + big_w + gap, big_y + 2, big_suffix, textAnchor="start",
                     fontName=font, fontSize=SUF_SIZE, fillColor=colors.HexColor("#94a3b8")))

    # badge — big_text 아래
    badge_bottom = big_y - 6
    if badge_text and badge_palette:
        bg, fg = badge_palette
        badge_w, badge_h = 66, 17
        badge_x = (W_pt - badge_w) / 2
        badge_y = badge_bottom - badge_h - 2
        d.add(Rect(badge_x, badge_y, badge_w, badge_h,
                   fillColor=colors.HexColor(bg), strokeColor=None, rx=8, ry=8))
        d.add(String(W_pt / 2, badge_y + 5, badge_text, textAnchor="middle",
                     fontName=font, fontSize=9, fillColor=colors.HexColor(fg)))

    # sub lines — 카드 하단
    if sub_line1:
        d.add(String(W_pt / 2, 18, sub_line1, textAnchor="middle",
                     fontName=font, fontSize=8.5, fillColor=colors.HexColor("#64748b")))
    if sub_line2:
        d.add(String(W_pt / 2, 8, sub_line2, textAnchor="middle",
                     fontName=font, fontSize=8, fillColor=colors.HexColor("#94a3b8")))
    return d


def _gradient_card_drawing(W_pt: float, H_pt: float, score: float, level: str,
                           confidence_pct: int, eval_text: str, font: str):
    """표지 좌측 그라데이션 카드 (Drawing). [사용처: §6.1 종합 점수]
    파랑→인디고 그라데이션 + 큰 점수 + 등급 + 신뢰도/평가 가능.
    """
    from reportlab.graphics.shapes import Drawing, Rect, String, Line
    from reportlab.lib import colors

    d = Drawing(W_pt, H_pt)
    # 그라데이션 흉내 — 가로 띠 6개
    grad_colors = ["#1e3a8a", "#1e40af", "#1d4ed8", "#2563eb", "#3b82f6", "#60a5fa"]
    strip_h = H_pt / len(grad_colors)
    for i, hex_c in enumerate(grad_colors):
        d.add(Rect(0, H_pt - (i + 1) * strip_h, W_pt, strip_h + 0.5,
                   fillColor=colors.HexColor(hex_c), strokeColor=None))
    # 외곽
    d.add(Rect(0, 0, W_pt, H_pt, fillColor=None, strokeColor=colors.HexColor("#1e40af"),
               strokeWidth=0.8))

    # 라벨
    d.add(String(W_pt / 2, H_pt - 22,
                 "종합 성숙도", textAnchor="middle",
                 fontName=font, fontSize=10, fillColor=colors.white))
    # 큰 점수
    score_text = f"{score:.2f}"
    d.add(String(W_pt / 2 - 20, H_pt / 2 - 6, score_text,
                 textAnchor="middle", fontName=font, fontSize=42, fillColor=colors.white))
    d.add(String(W_pt / 2 + 38, H_pt / 2 - 2, "/ 4.0",
                 textAnchor="middle", fontName=font, fontSize=12, fillColor=colors.HexColor("#dbeafe")))
    # 등급 박스
    badge_w, badge_h = 90, 22
    badge_x = (W_pt - badge_w) / 2
    badge_y = H_pt / 2 - 38
    d.add(Rect(badge_x, badge_y, badge_w, badge_h,
               fillColor=colors.white, strokeColor=None, rx=4, ry=4))
    d.add(String(W_pt / 2, badge_y + 7, f"{level} 단계",
                 textAnchor="middle", fontName=font, fontSize=11,
                 fillColor=colors.HexColor("#1e3a8a")))

    # 하단 신뢰도/평가 가능
    d.add(Line(10, 36, W_pt - 10, 36, strokeColor=colors.HexColor("#bfdbfe"), strokeWidth=0.5))
    d.add(String(W_pt / 4, 22, "진단 신뢰도", textAnchor="middle",
                 fontName=font, fontSize=8, fillColor=colors.HexColor("#dbeafe")))
    d.add(String(W_pt / 4, 10, f"{confidence_pct}%", textAnchor="middle",
                 fontName=font, fontSize=11, fillColor=colors.white))
    d.add(Line(W_pt / 2, 8, W_pt / 2, 30, strokeColor=colors.HexColor("#bfdbfe"), strokeWidth=0.5))
    d.add(String(W_pt * 3 / 4, 22, "평가 가능", textAnchor="middle",
                 fontName=font, fontSize=8, fillColor=colors.HexColor("#dbeafe")))
    d.add(String(W_pt * 3 / 4, 10, eval_text, textAnchor="middle",
                 fontName=font, fontSize=11, fillColor=colors.white))
    return d


def _radar_drawing(W_pt: float, H_pt: float, pillars: list[dict], font: str):
    """6각형 레이더 차트. 외곽=4.0."""
    from reportlab.graphics.shapes import Drawing, Polygon, Line, String, Circle
    from reportlab.lib import colors

    d = Drawing(W_pt, H_pt)
    cx, cy = W_pt / 2, H_pt / 2 - 4
    R = min(W_pt, H_pt) / 2 - 28

    n = max(len(pillars), 1)
    # 격자(4 단계)
    for ring in range(1, 5):
        rr = R * ring / 4
        pts = []
        for i in range(n):
            ang = math.pi / 2 - 2 * math.pi * i / n
            pts.extend([cx + rr * math.cos(ang), cy + rr * math.sin(ang)])
        d.add(Polygon(points=pts, fillColor=None,
                      strokeColor=colors.HexColor("#cbd5e1" if ring == 4 else "#e5e7eb"),
                      strokeWidth=0.4 if ring < 4 else 0.6))

    # 축 라인 + 축 라벨
    for i, p in enumerate(pillars):
        ang = math.pi / 2 - 2 * math.pi * i / n
        x_end = cx + R * math.cos(ang)
        y_end = cy + R * math.sin(ang)
        d.add(Line(cx, cy, x_end, y_end, strokeColor=colors.HexColor("#e5e7eb"),
                   strokeWidth=0.4))
        # 라벨 (반경 R+12 위치)
        lx = cx + (R + 14) * math.cos(ang)
        ly = cy + (R + 12) * math.sin(ang) - 4
        anchor = "middle"
        if math.cos(ang) > 0.3:
            anchor = "start"
        elif math.cos(ang) < -0.3:
            anchor = "end"
        d.add(String(lx, ly, f"{_ko_pillar_short(p['pillar'])} {p['score']:.2f}",
                     textAnchor=anchor, fontName=font, fontSize=8,
                     fillColor=colors.HexColor("#374151")))

    # 데이터 폴리곤
    pts = []
    for i, p in enumerate(pillars):
        ang = math.pi / 2 - 2 * math.pi * i / n
        ratio = max(0.0, min(1.0, (p["score"] or 0.0) / 4.0))
        rr = R * ratio
        pts.extend([cx + rr * math.cos(ang), cy + rr * math.sin(ang)])
    d.add(Polygon(points=pts,
                  fillColor=colors.Color(0.149, 0.388, 0.922, alpha=0.25),  # #2563eb @25%
                  strokeColor=colors.HexColor("#2563eb"),
                  strokeWidth=1.2))
    # 데이터 포인트
    for i, p in enumerate(pillars):
        ang = math.pi / 2 - 2 * math.pi * i / n
        ratio = max(0.0, min(1.0, (p["score"] or 0.0) / 4.0))
        rr = R * ratio
        d.add(Circle(cx + rr * math.cos(ang), cy + rr * math.sin(ang), 2.2,
                     fillColor=colors.HexColor("#2563eb"), strokeColor=colors.white,
                     strokeWidth=0.5))

    # 외곽 라벨
    d.add(String(W_pt - 4, 6, "외곽 = 4.0(최적화)", textAnchor="end",
                 fontName=font, fontSize=7, fillColor=colors.HexColor("#9ca3af")))
    return d


def _hbar_drawing(W_pt: float, pillars: list[dict], font: str):
    """필러별 막대 그래프."""
    from reportlab.graphics.shapes import Drawing, Rect, String, Line
    from reportlab.lib import colors

    bar_h, gap = 14, 8
    pad_top, pad_bottom = 8, 16
    H_pt = pad_top + pad_bottom + len(pillars) * (bar_h + gap)
    d = Drawing(W_pt, H_pt)

    label_w = 80
    score_w = 70
    bar_area_x = label_w
    bar_area_w = W_pt - label_w - score_w
    # 축 (4.0)
    d.add(Line(bar_area_x, pad_bottom - 2, bar_area_x + bar_area_w, pad_bottom - 2,
               strokeColor=colors.HexColor("#e5e7eb"), strokeWidth=0.4))
    for tick in (1, 2, 3, 4):
        tx = bar_area_x + bar_area_w * tick / 4
        d.add(Line(tx, pad_bottom - 4, tx, pad_bottom - 1,
                   strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.4))
        d.add(String(tx, pad_bottom - 12, str(tick), textAnchor="middle",
                     fontName=font, fontSize=6, fillColor=colors.HexColor("#9ca3af")))

    for i, p in enumerate(pillars):
        y = H_pt - pad_top - (i + 1) * (bar_h + gap) + gap
        score = max(0.0, min(4.0, p["score"] or 0.0))
        bw = bar_area_w * score / 4.0
        # bg
        d.add(Rect(bar_area_x, y, bar_area_w, bar_h,
                   fillColor=colors.HexColor("#f3f4f6"), strokeColor=None))
        # fg color by level
        col = colors.HexColor(_COLORS.get(p["level"], "#2563eb"))
        d.add(Rect(bar_area_x, y, bw, bar_h, fillColor=col, strokeColor=None))
        d.add(String(label_w - 6, y + bar_h / 2 - 3, _ko_pillar_short(p["pillar"]),
                     textAnchor="end", fontName=font, fontSize=8,
                     fillColor=colors.HexColor("#374151")))
        d.add(String(bar_area_x + bar_area_w + 6, y + bar_h / 2 - 3,
                     f"{score:.2f}  {p['level']}",
                     fontName=font, fontSize=8, fillColor=colors.HexColor("#374151")))
    return d


def _trend_drawing(W_pt: float, H_pt: float, history: list[dict], font: str):
    """종합 점수 추이 라인 차트."""
    from reportlab.graphics.shapes import Drawing, Line, String, Circle, Rect
    from reportlab.lib import colors

    d = Drawing(W_pt, H_pt)
    pad_l, pad_r, pad_t, pad_b = 30, 14, 14, 22
    x0, y0 = pad_l, pad_b
    w = W_pt - pad_l - pad_r
    h = H_pt - pad_t - pad_b

    # 축
    d.add(Line(x0, y0, x0, y0 + h, strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.5))
    d.add(Line(x0, y0, x0 + w, y0, strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.5))
    # Y 눈금
    for tick in (0, 1, 2, 3, 4):
        ty = y0 + h * tick / 4
        d.add(Line(x0 - 3, ty, x0, ty, strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=0.4))
        d.add(String(x0 - 6, ty - 3, str(tick), textAnchor="end",
                     fontName=font, fontSize=7, fillColor=colors.HexColor("#9ca3af")))
        if tick > 0:
            d.add(Line(x0, ty, x0 + w, ty, strokeColor=colors.HexColor("#f1f5f9"), strokeWidth=0.3))

    if not history:
        d.add(String(W_pt / 2, H_pt / 2, "추이 데이터 없음", textAnchor="middle",
                     fontName=font, fontSize=9, fillColor=colors.HexColor("#9ca3af")))
        return d

    n = len(history)
    pts = []
    for i, hp in enumerate(history):
        score = max(0.0, min(4.0, hp.get("total_score") or 0.0))
        x = x0 + (w * (i / max(n - 1, 1)) if n > 1 else w / 2)
        y = y0 + h * score / 4.0
        pts.append((x, y, hp))

    # 라인
    for i in range(len(pts) - 1):
        x1, y1, _ = pts[i]
        x2, y2, _ = pts[i + 1]
        d.add(Line(x1, y1, x2, y2, strokeColor=colors.HexColor("#2563eb"), strokeWidth=1.4))
    # 점 + 라벨
    for x, y, hp in pts:
        d.add(Circle(x, y, 3, fillColor=colors.HexColor("#2563eb"),
                     strokeColor=colors.white, strokeWidth=0.6))
        d.add(String(x, y + 6, f"{hp.get('total_score', 0):.2f}",
                     textAnchor="middle", fontName=font, fontSize=7,
                     fillColor=colors.HexColor("#1e3a5f")))
        # X 라벨 (YYYY-MM)
        at = (hp.get("assessed_at") or "")[:7]
        d.add(String(x, y0 - 12, at, textAnchor="middle",
                     fontName=font, fontSize=6.5, fillColor=colors.HexColor("#6b7280")))
    return d


def _make_pdf(data: dict) -> bytes:
    _ensure_font()
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (
        BaseDocTemplate, PageTemplate, Frame,
        Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak, KeepTogether,
    )

    F = _FONT_NAME
    PAGE_W, PAGE_H = A4
    MARGIN = 1.6 * cm
    CONTENT_W = PAGE_W - 2 * MARGIN

    s   = data["session"]
    sm  = data["summary"]
    ps  = data["pillar_scores"] or []
    cr  = data["checklist_results"] or []
    imps = data["improvements"] or []
    em  = data.get("evaluation_meta") or {}
    history = data.get("history") or []
    prev_h  = data.get("previous_session")
    std_sum = data.get("standards_summary") or {"nist_800_207": [], "cis_controls_v8": []}

    gen_iso = data["generated_at"]
    gen = gen_iso[:10]
    doc_no = _doc_no(s["session_id"], gen_iso)
    started = (s["started_at"] or "")[:10] or "-"
    completed = (s["completed_at"] or "")[:10] or "-"
    org_name = s["org"] or "(미상)"
    manager  = s["manager"] or "(미상)"

    overall_score = sm["overall_score"]
    overall_level = sm["overall_level"]
    total_items   = sm["total_items"]
    pass_cnt      = sm["pass_cnt"]
    partial_cnt   = sm["partial_cnt"]
    fail_cnt      = sm["fail_cnt"]
    na_cnt        = sm["na_cnt"]
    eval_capable  = total_items - na_cnt
    confidence_pct = int(round((sm.get("confidence") or 0.0) * 100))

    scan_mode = (em.get("scan_mode") or "demo").lower()
    scan_mode_label = {"demo": "데모", "live": "실 스캔"}.get(scan_mode, scan_mode)
    selected_tools = em.get("selected_tools") or []
    tool_matrix = " · ".join(
        {"keycloak": "Keycloak", "wazuh": "Wazuh", "nmap": "Nmap",
         "trivy": "Trivy", "web_probe": "web_probe",
         "supabase": "Supabase", "vercel": "Vercel", "railway": "Railway"}.get(t, t)
        for t in selected_tools
    ) or "(없음)"
    profile = em.get("profile_select") or {}
    idp  = profile.get("idp_type") or "none"
    siem = profile.get("siem_type") or "none"
    reviewers = em.get("reviewers") or {}

    # ── Styles ────────────────────────────────────────────────────────────
    def sty(name, **kw) -> ParagraphStyle:
        return ParagraphStyle(name, fontName=F, **kw)

    H_PART = sty("part_h", fontSize=18, leading=22,
                 textColor=colors.HexColor(_COLORS["navy"]), spaceAfter=4, spaceBefore=2)
    H_SEC  = sty("sec",    fontSize=13, leading=18,
                 textColor=colors.HexColor(_COLORS["navy"]), spaceAfter=4, spaceBefore=6)
    H_SUB  = sty("sub",    fontSize=10.5, leading=14,
                 textColor=colors.HexColor("#1f2937"), spaceAfter=2, spaceBefore=4)
    BODY   = sty("body",   fontSize=8.8, leading=12.5, textColor=colors.HexColor("#374151"))
    BODY_C = sty("body_c", fontSize=8.8, leading=12.5, textColor=colors.HexColor("#374151"),
                 alignment=1)
    SMALL  = sty("small",  fontSize=7.8, leading=10, textColor=colors.HexColor("#6b7280"))
    SMALL_R = sty("smallr", fontSize=7.8, leading=10, textColor=colors.HexColor("#6b7280"),
                  alignment=2)
    TINY   = sty("tiny",   fontSize=7,   leading=9,  textColor=colors.HexColor("#6b7280"))
    META_K = sty("meta_k", fontSize=8.5, leading=11, textColor=colors.HexColor("#6b7280"))
    META_V = sty("meta_v", fontSize=8.5, leading=11, textColor=colors.HexColor("#111827"))
    QUOTE  = sty("quote",  fontSize=8.5, leading=12, textColor=colors.HexColor("#475569"))
    CELL   = sty("cell",   fontSize=8,   leading=11, textColor=colors.HexColor("#374151"))
    CELL_B = sty("cell_b", fontSize=8,   leading=11, textColor=colors.HexColor("#111827"))

    # ── BaseDocTemplate (cover + body templates) ──────────────────────────
    buf = io.BytesIO()
    doc = BaseDocTemplate(
        buf, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN,
        title=f"제로트러스트 진단 보고서 - {org_name}",
        author="Readyz-T",
    )

    # PART 이름 추적
    current_part = {"name": "1부 · 문서 식별"}

    def _cover_page(canvas, doc_):
        canvas.saveState()
        # 상단 한 줄 영문
        canvas.setFont(F, 7)
        canvas.setFillColor(colors.HexColor("#94a3b8"))
        canvas.drawCentredString(PAGE_W / 2, PAGE_H - 1.4 * cm,
            "Z E R O T R U S T   M A T U R I T Y   A S S E S S M E N T   R E P O R T")
        # 영문 헤더 아래 얇은 가로선
        canvas.setStrokeColor(colors.HexColor("#e2e8f0"))
        canvas.setLineWidth(0.4)
        canvas.line(MARGIN, PAGE_H - 1.65 * cm, PAGE_W - MARGIN, PAGE_H - 1.65 * cm)
        # 푸터 (표지에는 안 그림 — 깔끔하게)
        canvas.restoreState()

    def _body_page(canvas, doc_):
        canvas.saveState()
        # 헤더
        canvas.setFont(F, 8)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(MARGIN, PAGE_H - 0.9 * cm, "제로트러스트 성숙도 진단 보고서")
        canvas.drawRightString(PAGE_W - MARGIN, PAGE_H - 0.9 * cm, current_part["name"])
        canvas.setStrokeColor(colors.HexColor("#e5e7eb"))
        canvas.setLineWidth(0.4)
        canvas.line(MARGIN, PAGE_H - 1.05 * cm, PAGE_W - MARGIN, PAGE_H - 1.05 * cm)
        # 푸터
        canvas.setFont(F, 7.5)
        canvas.setFillColor(colors.HexColor("#9ca3af"))
        canvas.drawString(MARGIN, 0.9 * cm, f"{doc_no} · v1.0")
        canvas.drawRightString(PAGE_W - MARGIN, 0.9 * cm, str(doc_.page))
        canvas.restoreState()

    cover_frame = Frame(MARGIN, MARGIN, CONTENT_W, PAGE_H - 2 * MARGIN,
                        id="cover_f", leftPadding=0, rightPadding=0,
                        topPadding=0, bottomPadding=0)
    body_frame  = Frame(MARGIN, MARGIN + 0.4 * cm, CONTENT_W,
                        PAGE_H - 2 * MARGIN - 0.7 * cm,
                        id="body_f", leftPadding=0, rightPadding=0,
                        topPadding=0.4 * cm, bottomPadding=0.2 * cm)

    doc.addPageTemplates([
        PageTemplate(id="cover", frames=[cover_frame], onPage=_cover_page),
        PageTemplate(id="body",  frames=[body_frame],  onPage=_body_page),
    ])

    story = []

    # =====================================================================
    # 1부 · 문서 식별
    # =====================================================================

    # ── 0. 표지 (샘플 PDF 픽셀 매칭) ─────────────────────────────────────
    # 상단 영문 라벨 (자간 넓게) — cover 페이지 헤더에서 그림 (onPage)
    story.append(Spacer(1, 2.4 * cm))

    story.append(Paragraph(
        '<font size="12" color="#64748b">제로트러스트 가이드라인 2.0 기반</font>',
        sty("cover_title1", fontSize=12, leading=18, alignment=1),
    ))
    story.append(Spacer(1, 0.1 * cm))
    story.append(Paragraph(
        '<font size="26" color="#0f172a"><b>보안 성숙도 진단 보고서</b></font>',
        sty("cover_title2", fontSize=26, leading=34, alignment=1),
    ))
    story.append(Spacer(1, 0.1 * cm))
    story.append(Paragraph(
        '<font size="10" color="#94a3b8"><i>Readyz-T Zero Trust Security Assessment</i></font>',
        sty("cover_title3", fontSize=10, leading=14, alignment=1),
    ))
    story.append(Spacer(1, 1.6 * cm))

    # 카드 3개 — 종합 성숙도 / 진단 신뢰도 / 진단 모드 (가로 배치)
    card_w = (CONTENT_W - 1.0 * cm) / 3  # 두 칸 사이 0.5cm 간격
    card_h = 3.4 * cm
    level_palette = _LEVEL_BADGE_COLORS.get(overall_level, ("#dbeafe", "#1e40af"))

    card_score = _cover_stat_card(
        card_w, card_h, "종합 성숙도", f"{overall_score:.2f}",
        big_color="#0f172a",
        badge_text=f"{overall_level} 단계", badge_palette=level_palette,
        sub_line1="", sub_line2="", font=F,
        big_suffix="/ 4.0",
    )
    card_conf = _cover_stat_card(
        card_w, card_h, "진단 신뢰도", f"{confidence_pct}%",
        big_color="#1e40af",
        badge_text=None, badge_palette=None,
        sub_line1=f"평가 가능 {eval_capable} / {total_items}", sub_line2="", font=F,
    )
    # 진단 모드 + 도구 매트릭스
    card_mode = _cover_stat_card(
        card_w, card_h, "진단 모드", scan_mode_label,
        big_color="#0f172a",
        badge_text=None, badge_palette=None,
        sub_line1=tool_matrix, sub_line2="", font=F,
    )

    cards_row = Table(
        [[card_score, card_conf, card_mode]],
        colWidths=[card_w, card_w, card_w],
        style=TableStyle([
            ("VALIGN",       (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING",   (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
            ("LEFTPADDING",  (1, 0), (1, 0), 0.5 * cm),
            ("LEFTPADDING",  (2, 0), (2, 0), 0.5 * cm),
        ]),
    )
    story.append(cards_row)
    story.append(Spacer(1, 1.1 * cm))

    # 메타 표 6행 (헤더 행 없음, 좌측 라벨 + 우측 값)
    meta_rows = [
        [Paragraph("진단 대상", META_K), Paragraph(org_name, META_V)],
        [Paragraph("담당자", META_K),    Paragraph(manager, META_V)],
        [Paragraph("진단 기간", META_K), Paragraph(f"{started} ~ {completed}", META_V)],
        [Paragraph("보고서 생성일", META_K), Paragraph(gen, META_V)],
        [Paragraph("문서번호 / 버전", META_K),
         Paragraph(f"{doc_no} / v1.0", META_V)],
        [Paragraph("기밀 등급", META_K),
         Paragraph('<font color="#b91c1c"><b>대외비 (Confidential)</b></font>', META_V)],
    ]
    meta_table = Table(
        meta_rows, colWidths=[3.4 * cm, CONTENT_W - 3.4 * cm],
        style=TableStyle([
            ("FONTNAME",      (0, 0), (-1, -1), F),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("BOX",           (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("INNERGRID",     (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
            ("BACKGROUND",    (0, 0), (0, -1), colors.HexColor("#f8fafc")),
            ("TOPPADDING",    (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LEFTPADDING",   (0, 0), (-1, -1), 12),
        ]),
    )
    story.append(meta_table)

    # 푸터 (페이지 맨 아래) — 빈 공간 채우고 가운데 텍스트
    story.append(Spacer(1, 3.5 * cm))
    story.append(Paragraph(
        f'<font size="9" color="#64748b">Readyz-T · 신뢰많이된다 팀 · {org_name}</font>',
        sty("cover_footer", fontSize=9, leading=12, alignment=1),
    ))

    # 본문 페이지 템플릿으로 전환
    story.append(PageBreak())
    from reportlab.platypus.doctemplate import NextPageTemplate
    # PageBreak 이후 새 페이지부터 body 적용
    story.insert(len(story) - 1, NextPageTemplate("body"))

    # =====================================================================
    # 1. 문서 정보
    # =====================================================================
    def _set_part(name: str):
        story.append(_PartMarker(name, current_part))

    _set_part("1부 · 문서 식별")
    story.append(_part_badge_flowable("1부 · 문서 식별", F))
    story.append(Spacer(1, 0.4 * cm))

    story.append(_section_heading_flowable("1", "문서 정보", F, CONTENT_W))
    story.append(Paragraph(
        "본 보고서가 결재·감사·대외 제출에 쓰일 수 있도록, "
        "문서로서의 식별 정보와 책임 소재를 명시합니다.", BODY))
    story.append(Spacer(1, 0.25 * cm))

    # 1.1 결재란
    story.append(Paragraph("1.1 결재란", H_SUB))
    approval_rows = [
        ["구분", "성명 / 서명", "직책", "일자"],
        ["판정자(작성)", manager, reviewers.get("app_owner") or "", gen],
        ["검토자",      reviewers.get("backend_owner") or "", "", ""],
        ["승인자",      reviewers.get("security_reviewer") or "", "", ""],
    ]
    story.append(_table(approval_rows, F,
                        col_widths=[3 * cm, 6 * cm, 4 * cm, CONTENT_W - 13 * cm],
                        header=True))
    story.append(Spacer(1, 0.3 * cm))

    # 1.2 문서 개정 이력
    story.append(Paragraph("1.2 문서 개정 이력", H_SUB))
    revision_rows = [
        ["버전", "일자", "변경 요약", "작성"],
        ["v1.0", gen, f"최초 발행 ({started} 진단 세션 기준)", manager],
    ]
    story.append(_table(revision_rows, F,
                        col_widths=[2 * cm, 2.5 * cm, CONTENT_W - 7.5 * cm, 3 * cm],
                        header=True))
    story.append(Spacer(1, 0.3 * cm))

    # 1.3 배포 및 열람 범위
    story.append(Paragraph("1.3 배포 및 열람 범위", H_SUB))
    dist_rows = [
        ["기밀 등급", "대외비 (Confidential) — 조직 외부 공유 금지"],
        ["열람 권한", "진단 대상 조직 보안 담당자, 경영진, 지정 감사인"],
        ["배포 형식", "PDF (본 문서) / 대시보드 공유 링크(읽기 전용, 기본 7일)"],
    ]
    story.append(_kv_table(dist_rows, F, key_w=3.2 * cm, total_w=CONTENT_W))
    story.append(Spacer(1, 0.3 * cm))

    # 1.4 스냅샷 고지
    story.append(Paragraph("1.4 스냅샷 고지", H_SUB))
    story.append(_callout_box(
        f"본 문서는 <b>{started}</b> 진단 세션 시점의 스냅샷입니다. "
        f"진단 이후 보안 구성이 변경되었을 수 있으며, 최신 현황·실시간 추이·"
        f"세션 간 비교는 ReadyzT 대시보드에서 확인하십시오. "
        f"본 문서의 점수와 판정은 진단 시점에 고정됩니다.",
        BODY, F, CONTENT_W, bg="#fffbeb", border="#fcd34d",
    ))
    story.append(Spacer(1, 0.4 * cm))

    # 2. 목차
    story.append(_section_heading_flowable("2", "목차", F, CONTENT_W))
    toc_rows = [
        ["부", "장"],
        ["1부 문서 식별",     "0. 표지 / 1. 문서 정보 / 2. 목차"],
        ["2부 요약",          "3. 경영진 요약"],
        ["3부 진단 개요",     "4. 진단 범위 및 방법론 / 5. 점수 산정 방식"],
        ["4부 종합 결과",     "6. 종합 성숙도 / 7. 목표 갭 분석 / 8. 추이·비교"],
        ["5부 상세 결과",     "9. 필러별 상세 / 10. 체크리스트 세부 / 11. 미충족·평가불가 정리"],
        ["6부 개선",          "12. 개선 로드맵 (30/60/90일)"],
        ["7부 부록",          "13. 표준 매핑 / 14. 증적 레지스터 / 15. 판정 로그 / 16. OCSF / 17. 용어집"],
    ]
    story.append(_table(toc_rows, F,
                        col_widths=[4 * cm, CONTENT_W - 4 * cm],
                        header=True))

    _set_part("2부 · 요약")
    story.append(PageBreak())

    # =====================================================================
    # 2부 · 요약 — 3. 경영진 요약
    # =====================================================================
    story.append(_part_badge_flowable("2부 · 요약", F))
    story.append(Spacer(1, 0.4 * cm))

    story.append(_section_heading_flowable("3", "경영진 요약", F, CONTENT_W))
    story.append(Paragraph(
        "의사결정자가 이 페이지만으로 현재 수준과 우선 과제를 파악할 수 있도록 핵심만 정리했습니다.",
        BODY))
    story.append(Spacer(1, 0.2 * cm))

    # 총평 자동 생성
    sorted_pillars_asc = sorted(ps, key=lambda x: x["score"])
    weak3 = sorted_pillars_asc[:3]
    weak_names = " · ".join(_ko_pillar_short(p["pillar"]) for p in weak3[:2]) or "(데이터 없음)"
    auto_summary = (
        f"<b>총평.</b> {org_name}은(는) 6개 필러 평균 "
        f"<b>{overall_score:.2f}점('{overall_level}' 단계)</b>으로, 현재 가장 취약한 영역은 "
        f"<b>{weak_names}</b>입니다. "
        + ("전반적인 통제 정착이 우선 과제이며, 단기적으로는 설정 변경 위주의 Quick Win 부터 착수할 것을 권고합니다."
           if overall_score < 2.5 else
           "통제 기반은 갖추어졌으나 자동화·실시간 대응·연동 영역의 보완이 다음 단계의 핵심입니다.")
    )
    story.append(_callout_box(auto_summary, BODY, F, CONTENT_W,
                              bg="#eff6ff", border="#93c5fd"))
    story.append(Spacer(1, 0.35 * cm))

    # 3.1 핵심 지표 4 카드
    story.append(Paragraph("3.1 핵심 지표", H_SUB))
    card_data = [[
        _stat_cell(pass_cnt, "충족", "#16a34a", "#dcfce7", F),
        _stat_cell(partial_cnt, "부분충족", "#92400e", "#fef9c3", F),
        _stat_cell(fail_cnt, "미충족", "#b91c1c", "#fee2e2", F),
        _stat_cell(na_cnt, "평가불가", "#475569", "#f1f5f9", F),
        _stat_cell(len(imps), "개선 과제", "#1e40af", "#dbeafe", F),
    ]]
    cw = CONTENT_W / 5
    story.append(Table(card_data, colWidths=[cw] * 5,
        style=TableStyle([
            ("VALIGN",       (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING",   (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
        ])))
    story.append(Spacer(1, 0.4 * cm))

    # 3.2 가장 취약한 3개 필러
    story.append(Paragraph("3.2 가장 취약한 3개 필러", H_SUB))
    weak_reasons = _weak_pillar_reasons(weak3, cr)
    weak_rows = [["순위", "필러", "점수", "등급", "주요 미흡 사유"]]
    for i, p in enumerate(weak3, start=1):
        weak_rows.append([
            str(i),
            _ko_pillar_short(p["pillar"]),
            f"{p['score']:.2f}",
            p["level"],
            Paragraph(weak_reasons.get(p["pillar"], "(미흡 사유 데이터 없음)"), CELL),
        ])
    story.append(_table(weak_rows, F,
        col_widths=[1.2 * cm, 3.0 * cm, 1.6 * cm, 2.0 * cm,
                    CONTENT_W - 7.8 * cm],
        header=True, level_col=3))
    story.append(Spacer(1, 0.4 * cm))

    # 3.3 Quick Win Top 3
    story.append(Paragraph("3.3 즉시 조치 권고 — Quick Win Top 3", H_SUB))
    story.append(Paragraph(
        "위험 대비 투입 노력이 가장 효율적인(빠른 개선) 항목입니다. "
        "설정 변경 위주로 단기 내 처리가 가능합니다.", SMALL))
    story.append(Spacer(1, 0.1 * cm))
    quick_wins = _quick_wins(imps, 3)
    qw_rows = [["#", "과제", "도구", "실행 요지"]]
    if not quick_wins:
        qw_rows.append(["—", "(개선 권고 데이터 없음)", "—", "—"])
    else:
        for i, q in enumerate(quick_wins, start=1):
            tool_disp = {"keycloak": "Keycloak", "wazuh": "Wazuh", "nmap": "Nmap",
                         "trivy": "Trivy", "web_probe": "web_probe",
                         "supabase": "Supabase", "vercel": "Vercel", "railway": "Railway",
                         "수동": "수동"}.get(
                (q.get("tool") or "").lower(), q.get("tool") or "—")
            # Paragraph 로 wrap 가능하게 (긴 텍스트 한 줄 잘림 방지)
            task_text = (q.get("task") or "")[:120]
            ex_text   = (q.get("solution") or q.get("expected_effect") or "")[:200]
            qw_rows.append([
                str(i),
                Paragraph(task_text, CELL),
                tool_disp,
                Paragraph(ex_text, CELL),
            ])
    from reportlab.platypus import KeepTogether
    _qw_table = _table(qw_rows, F,
        col_widths=[0.8 * cm, 4.5 * cm, 2.2 * cm, CONTENT_W - 7.5 * cm],
        header=True)
    # 표 + ※ 안내 문구를 한 묶음으로 — 안내가 단독으로 빈 페이지에 떨어지지 않게.
    story.append(KeepTogether([
        _qw_table,
        Spacer(1, 0.3 * cm),
        Paragraph(
            "※ 상세 근거는 5부(상세 결과), 전체 실행 계획은 6부(개선 로드맵)를 참조하십시오.",
            SMALL),
    ]))

    _set_part("3부 · 진단 개요")
    story.append(PageBreak())

    # =====================================================================
    # 3부 · 진단 개요
    # =====================================================================
    story.append(_part_badge_flowable("3부 · 진단 개요", F))
    story.append(Spacer(1, 0.4 * cm))

    story.append(_section_heading_flowable("4", "진단 범위 및 방법론", F, CONTENT_W))
    story.append(Paragraph(
        "제로트러스트 가이드라인 2.0의 평가 확정사항(가이드 §3)에 따른 진단 전제와 범위입니다.",
        BODY))
    story.append(Spacer(1, 0.25 * cm))

    # 4.1 평가 기준 및 범위
    story.append(Paragraph("4.1 평가 기준 및 범위", H_SUB))
    krit_rows = [
        ["평가 기준",    "제로트러스트 가이드라인 2.0 · 6개 필러 체계"],
        ["평가 항목 수", f"총 {total_items}개 (필러별 세부 단계 항목)"],
        ["대상 자산",
         f"{org_name} 진단 범위 내 IT 자산 ({'데모 환경' if scan_mode == 'demo' else '실 운영 환경'})"],
        ["데이터 등급", "대외비 / 일반 — 진단 결과 산출물에 한함"],
        ["판정자",      manager],
    ]
    story.append(_kv_table(krit_rows, F, key_w=3.2 * cm, total_w=CONTENT_W))
    story.append(Spacer(1, 0.3 * cm))

    # 4.2 진단 환경 및 도구 분담
    story.append(Paragraph("4.2 진단 환경 및 도구 분담", H_SUB))
    tool_count_map = _tool_auto_counts(cr)
    tool_table_rows = [
        ["도구", "대상 필러", "자동 평가", "수집 범위"],
        ["Keycloak", "식별자·신원", str(tool_count_map.get("keycloak", 0)),
         "사용자 인벤토리·MFA·세션·RBAC/ABAC·감사로깅"],
        ["Wazuh", "기기·시스템·데이터", str(tool_count_map.get("wazuh", 0)),
         "에이전트·SCA·FIM·행위탐지·자동대응"],
        ["Nmap", "네트워크", str(tool_count_map.get("nmap", 0)),
         "호스트·포트·서브넷·TLS·VPN"],
        ["Trivy", "앱·워크로드", str(tool_count_map.get("trivy", 0)),
         "이미지·파일시스템 취약점·SBOM·의존성"],
        ["web_probe", "앱·데이터", str(tool_count_map.get("web_probe", 0)),
         "공개 URL 보안 헤더·HTTPS·DNS 위생"],
        ["Supabase", "식별자·신원·데이터", str(tool_count_map.get("supabase", 0)),
         "Auth 정책·RLS·사용자·MFA·비밀번호 정책"],
        ["Vercel", "앱·데이터", str(tool_count_map.get("vercel", 0)),
         "배포 이력·환경변수 분리·도메인 SSL·팀 RBAC"],
        ["Railway", "네트워크·앱", str(tool_count_map.get("railway", 0)),
         "배포 상태·환경변수·헬스체크·restart 정책"],
        ["수동", "전 영역", str(tool_count_map.get("수동", 0)),
         "도구 미연결·정책/문서 기반 항목 (폴백 자동 전환)"],
    ]
    story.append(_table(tool_table_rows, F,
        col_widths=[2.2 * cm, 3.4 * cm, 1.8 * cm, CONTENT_W - 7.4 * cm],
        header=True))
    story.append(Spacer(1, 0.15 * cm))
    story.append(Paragraph(
        f"※ 진단 모드: <b>{scan_mode_label}</b>"
        + (" (외부 시스템 미접근, 결정론적 가상 결과)" if scan_mode == "demo" else " (외부 시스템 실제 접근)")
        + f". IdP: <b>{idp}</b> / SIEM: <b>{siem}</b>.",
        SMALL))
    story.append(Spacer(1, 0.3 * cm))

    # 4.3 진단 일정
    story.append(Paragraph("4.3 진단 일정", H_SUB))
    sch_rows = [
        ["진단 시작",   started],
        ["진단 완료",   completed],
        ["보고서 생성", gen],
    ]
    story.append(_kv_table(sch_rows, F, key_w=3.2 * cm, total_w=CONTENT_W))
    story.append(Spacer(1, 0.4 * cm))

    # 5. 점수 산정 방식
    story.append(_section_heading_flowable("5", "점수 산정 방식", F, CONTENT_W))
    story.append(Paragraph(
        "점수의 의미와 한계를 독자가 스스로 해석할 수 있도록 산정 로직을 명시합니다.", BODY))
    story.append(Spacer(1, 0.2 * cm))

    # 5.1 항목 판정 기준
    story.append(Paragraph("5.1 항목 판정 기준", H_SUB))
    crit_rows = [
        ["판정", "조건 (측정값 vs 임계값)", "가중치", "점수 반영"],
        ["충족",     "측정값 ≥ 임계값",                    "1.0", "전체 반영"],
        ["부분충족", "임계값 × 0.7 ≤ 측정값 < 임계값",     "0.5", "절반 반영"],
        ["미충족",   "측정값 < 임계값 × 0.7",               "0.0", "미반영"],
        ["평가불가", "오류 / 도구 미연결",                  "—",   "분모에서 제외"],
    ]
    story.append(_table(crit_rows, F,
        col_widths=[2.0 * cm, CONTENT_W - 6.4 * cm, 1.8 * cm, 2.6 * cm],
        header=True, result_col=0))
    story.append(Spacer(1, 0.1 * cm))
    story.append(Paragraph(
        "※ 역방향 항목(임계값=0, 낮을수록 좋음): 0 → 충족 / ≤5 → 부분충족 / &gt;5 → 미충족",
        SMALL))
    story.append(Spacer(1, 0.3 * cm))

    # 5.2 필러 점수 산식
    story.append(Paragraph("5.2 필러 점수 산식", H_SUB))
    story.append(_callout_box(
        "<b>필러 점수 =</b> ( Σ(성숙도 × 가중치) ÷ Σ(성숙도) ) × 4.0<br/>"
        "<font color='#6b7280' size='8'>평가불가 항목은 분모·분자 모두에서 제외됩니다.</font>",
        BODY, F, CONTENT_W, bg="#f8fafc", border="#cbd5e1"))
    story.append(Spacer(1, 0.3 * cm))

    # 5.3 성숙도 4단계
    story.append(Paragraph("5.3 성숙도 4단계", H_SUB))
    def _lvl_chip(name: str):
        bg, fg = _LEVEL_BADGE_COLORS.get(name, ("#e5e7eb", "#475569"))
        chip_sty = ParagraphStyle(
            f"_lvl_{name}", fontName=F, fontSize=9, leading=12, alignment=1,
            textColor=colors.HexColor(fg), backColor=colors.HexColor(bg),
            borderPadding=(3, 8, 3, 8), borderRadius=6,
        )
        return Paragraph(f"<b>{name}</b>", chip_sty)

    mat_rows = [
        ["단계", "점수 구간"],
        [_lvl_chip("최적화"), "3.5 이상"],
        [_lvl_chip("향상"),   "2.5 이상 ~ 3.5 미만"],
        [_lvl_chip("초기"),   "1.5 이상 ~ 2.5 미만"],
        [_lvl_chip("기존"),   "1.5 미만"],
    ]
    story.append(_table(mat_rows, F,
        col_widths=[3 * cm, CONTENT_W - 3 * cm],
        header=True))
    story.append(Spacer(1, 0.15 * cm))
    next_level, gap = _next_level_gap(overall_score)
    story.append(Paragraph(
        f"현재 종합 <b>{overall_score:.2f}점</b>은 '<b>{overall_level}</b>' 구간이며, "
        + (f"'<b>{next_level}</b>' 단계까지 <b>{gap:.2f}점</b>이 부족합니다."
           if next_level else "이미 최고 단계('최적화')입니다."),
        BODY))
    story.append(Spacer(1, 0.3 * cm))

    # 5.4 신뢰도
    story.append(Paragraph("5.4 신뢰도(confidence)", H_SUB))
    story.append(_callout_box(
        "<b>confidence =</b> 평가 가능 항목 ÷ 전체 항목 × 100%<br/>"
        f"<b>본 진단:</b> {eval_capable} / {total_items} = "
        f"<b>{confidence_pct}%</b> "
        f"({scan_mode_label} 모드, 평가불가 {na_cnt}건)",
        BODY, F, CONTENT_W, bg="#f8fafc", border="#cbd5e1"))

    _set_part("4부 · 종합 결과")
    story.append(PageBreak())

    # =====================================================================
    # 4부 · 종합 결과
    # =====================================================================
    story.append(_part_badge_flowable("4부 · 종합 결과", F))
    story.append(Spacer(1, 0.4 * cm))

    story.append(_section_heading_flowable("6", "종합 성숙도 분석", F, CONTENT_W))
    story.append(Paragraph(
        "대시보드 종합 화면을 문서로 고정한 것으로, 6개 필러의 성숙도를 한눈에 비교합니다.",
        BODY))
    story.append(Spacer(1, 0.3 * cm))

    # 6.1 종합 점수 카드 + 6.2 레이더 (2단)
    score_card_w = 5.0 * cm
    score_card_h = 4.6 * cm
    radar_w = CONTENT_W - score_card_w - 0.4 * cm
    radar_h = 6.4 * cm
    _eval_text_local = f"{eval_capable} / {total_items}"
    score_card = _gradient_card_drawing(score_card_w, score_card_h, overall_score,
                                        overall_level, confidence_pct, _eval_text_local, F)
    radar = _radar_drawing(radar_w, radar_h, ps, F) if ps else Paragraph(
        "(필러 데이터 없음)", SMALL)

    sec_a = Table(
        [[Paragraph("6.1 종합 점수", H_SUB)],
         [score_card]],
        colWidths=[score_card_w],
        style=TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0),
                          ("RIGHTPADDING", (0, 0), (-1, -1), 0)]),
    )
    sec_b = Table(
        [[Paragraph("6.2 필러별 레이더 차트", H_SUB)],
         [radar]],
        colWidths=[radar_w],
        style=TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0),
                          ("RIGHTPADDING", (0, 0), (-1, -1), 0)]),
    )
    story.append(Table(
        [[sec_a, sec_b]],
        colWidths=[score_card_w, radar_w + 0.4 * cm],
        style=TableStyle([
            ("VALIGN",       (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("LEFTPADDING",  (1, 0), (1, 0), 8),
        ]),
    ))
    story.append(Spacer(1, 0.4 * cm))

    # 6.3 필러별 막대
    story.append(Paragraph("6.3 필러별 막대 그래프", H_SUB))
    if ps:
        story.append(_hbar_drawing(CONTENT_W, ps, F))
    else:
        story.append(Paragraph("(필러 데이터 없음)", SMALL))
    story.append(Spacer(1, 0.4 * cm))

    # 6.4 판정 분포
    story.append(Paragraph("6.4 판정 분포", H_SUB))
    dist_rows = [["필러", "점수", "등급", "충족", "부분충족", "미충족", "평가불가"]]
    sums = [0, 0, 0, 0]
    for p in ps:
        # pillar별 부분충족 계산 (MaturityScore 에는 partial_cnt 가 없음 — checklist_results 에서 집계)
        partial = sum(1 for r in cr
                      if r["pillar"] == p["pillar"] and r["result"] == "부분충족")
        dist_rows.append([
            _ko_pillar_short(p["pillar"]),
            f"{p['score']:.2f}",
            p["level"],
            str(p["pass_cnt"]),
            str(partial),
            str(p["fail_cnt"]),
            str(p["na_cnt"]),
        ])
        sums[0] += p["pass_cnt"]
        sums[1] += partial
        sums[2] += p["fail_cnt"]
        sums[3] += p["na_cnt"]
    dist_rows.append(["합계", f"{overall_score:.2f}", overall_level,
                      str(sums[0]), str(sums[1]), str(sums[2]), str(sums[3])])
    story.append(_table(dist_rows, F,
        col_widths=[3.4 * cm, 1.6 * cm, 1.8 * cm,
                    (CONTENT_W - 6.8 * cm) / 4,
                    (CONTENT_W - 6.8 * cm) / 4,
                    (CONTENT_W - 6.8 * cm) / 4,
                    (CONTENT_W - 6.8 * cm) / 4],
        header=True, level_col=2, sum_row=True))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(
        "※ 부분충족은 단계별 집계 특성상 합계행에 별도 표기. "
        "필러별 충족/실패 분류는 단계 충족 기준입니다.", SMALL))
    story.append(PageBreak())

    # 7. 목표 갭 분석
    story.append(_section_heading_flowable("7", "목표 대비 갭 분석", F, CONTENT_W))
    story.append(Paragraph(
        "조직이 설정한 목표 성숙도(전체 3.0 가정)와 현재 점수의 격차를 큰 순으로 정렬했습니다.",
        BODY))
    story.append(Spacer(1, 0.2 * cm))

    target = 3.0
    sorted_gap = sorted(ps, key=lambda x: x["score"])
    gap_rows = [["우선", "필러", "현재", "목표", "갭"]]
    for i, p in enumerate(sorted_gap, start=1):
        gap_rows.append([
            str(i),
            _ko_pillar_short(p["pillar"]),
            f"{p['score']:.2f}",
            f"{target:.1f}",
            f"{p['score'] - target:+.2f}",
        ])
    story.append(_table(gap_rows, F,
        col_widths=[1.2 * cm, 4.0 * cm, 1.8 * cm, 1.8 * cm,
                    CONTENT_W - 8.8 * cm],
        header=True))
    if sorted_gap:
        top3 = " → ".join(_ko_pillar_short(p["pillar"]) for p in sorted_gap[:3])
        story.append(Spacer(1, 0.15 * cm))
        story.append(Paragraph(f"우선 개선 영역은 <b>{top3}</b> 순입니다. "
                               "목표 수준은 조직 설정값(7부 부록 또는 설정 화면)에 따라 달라집니다.",
                               BODY))
    story.append(Spacer(1, 0.4 * cm))

    # 8. 추이 / 이전 대비 비교
    story.append(_section_heading_flowable("8", "진단 추이 / 이전 대비 비교", F, CONTENT_W))
    story.append(Paragraph(
        "시간순 점수 변화와 직전 세션 대비 항목별 변화를 문서로 고정합니다.", BODY))
    story.append(Spacer(1, 0.2 * cm))

    story.append(Paragraph("8.1 종합 점수 추이", H_SUB))
    if len(history) >= 1:
        story.append(_trend_drawing(CONTENT_W, 4.6 * cm, history, F))
    else:
        story.append(Paragraph("(추이 데이터 없음)", SMALL))
    story.append(Spacer(1, 0.3 * cm))

    story.append(Paragraph("8.2 직전 세션 대비 변화", H_SUB))
    if prev_h is None:
        story.append(_callout_box(
            "<b>기준선(baseline) 진단</b><br/>"
            "본 진단은 해당 조직의 첫 진단(기준선)으로, 비교 대상 이전 세션이 없습니다. "
            "차기 진단부터 필러별 ▲/▼ 변화와 항목별 delta(개선/악화)가 이 섹션에 표시됩니다.",
            BODY, F, CONTENT_W, bg="#f0f9ff", border="#7dd3fc"))
        diff_rows = [["필러", "이번", "직전", "Δ"]]
        diff_rows.append(["전체", f"{overall_score:.2f}", "—", "신규"])
    else:
        diff_rows = [["필러", "이번", "직전", "Δ"]]
        prev_pillars = prev_h.get("pillar_scores") or {}
        for p in ps:
            prev_score = prev_pillars.get(p["pillar"])
            if prev_score is None:
                diff_rows.append([_ko_pillar_short(p["pillar"]),
                                  f"{p['score']:.2f}", "—", "신규"])
            else:
                delta = p["score"] - prev_score
                arrow = "▲" if delta > 0 else ("▼" if delta < 0 else "—")
                diff_rows.append([_ko_pillar_short(p["pillar"]),
                                  f"{p['score']:.2f}",
                                  f"{prev_score:.2f}",
                                  f"{arrow} {delta:+.2f}"])
        diff_rows.append(["전체", f"{overall_score:.2f}",
                          f"{prev_h.get('total_score', 0):.2f}",
                          f"{overall_score - (prev_h.get('total_score') or 0):+.2f}"])
    story.append(_table(diff_rows, F,
        col_widths=[CONTENT_W - 7.8 * cm, 2.6 * cm, 2.6 * cm, 2.6 * cm],
        header=True, sum_row=True if prev_h else False))

    _set_part("5부 · 상세 결과")
    story.append(PageBreak())

    # =====================================================================
    # 5부 · 상세 결과 — 9. 필러별 상세
    # =====================================================================
    story.append(_part_badge_flowable("5부 · 상세 결과", F))
    story.append(Spacer(1, 0.4 * cm))

    story.append(_section_heading_flowable("9", "필러별 상세", F, CONTENT_W))
    story.append(Paragraph(
        "필러마다 ① 요약 박스 → ② 카테고리·단계별 항목 순으로 구성됩니다.", BODY))
    story.append(Spacer(1, 0.25 * cm))

    # 필러 순서: pillar_scores 순 (DB 저장 순)
    by_pillar: dict[str, list] = {}
    for it in cr:
        by_pillar.setdefault(it["pillar"], []).append(it)

    MATURITY_ORDER = {"기존": 1, "초기": 2, "향상": 3, "최적화": 4}
    for pi, p in enumerate(ps, start=1):
        pillar = p["pillar"]
        items = by_pillar.get(pillar, [])
        pass_n = sum(1 for x in items if x["result"] == "충족")
        partial_n = sum(1 for x in items if x["result"] == "부분충족")
        fail_n = sum(1 for x in items if x["result"] == "미충족")
        na_n = sum(1 for x in items if x["result"] == "평가불가")
        urgent = _urgent_item(items)
        # 9.x.1 요약 박스
        story.append(Paragraph(f"9.{pi} {pillar}", H_SUB))
        summary_html = (
            f"<b>충족 {pass_n}</b> · 부분충족 {partial_n} · "
            f"<font color='#b91c1c'>미충족 {fail_n}</font> · "
            f"평가불가 {na_n}  &nbsp;&nbsp; "
            f"점수 <b>{p['score']:.2f}</b> ({p['level']})<br/>"
            f"<font color='#6b7280'>가장 시급: {urgent}</font>"
        )
        story.append(_callout_box(summary_html, BODY, F, CONTENT_W,
                                  bg="#f8fafc", border="#cbd5e1"))
        story.append(Spacer(1, 0.2 * cm))

        # 카테고리별
        by_cat: dict[str, list] = {}
        for it in items:
            by_cat.setdefault(it["category"] or "(미분류)", []).append(it)

        for cat_idx, (cat, cit) in enumerate(sorted(by_cat.items()), start=1):
            sorted_items = sorted(cit, key=lambda x: MATURITY_ORDER.get(x.get("maturity"), 5))
            n_stage = len(sorted_items)
            n_pass = sum(1 for x in sorted_items if x["result"] == "충족")
            # rep stage
            highest_pass = [x for x in sorted_items if x["result"] == "충족"]
            rep_mat = highest_pass[-1]["maturity"] if highest_pass else (
                sorted_items[0]["maturity"] if sorted_items else "-")
            story.append(Paragraph(
                f"<b>{cat}</b> &nbsp;&nbsp;"
                f"<font color='#6b7280' size='8'>(현재 단계: "
                f"<b>{rep_mat}</b> · {n_pass}/{n_stage} 단계 충족)</font>",
                sty("cat_h", fontSize=9.5, leading=12,
                    textColor=colors.HexColor("#1f2937"), spaceBefore=3, spaceAfter=3),
            ))
            # 단계별 표
            cat_rows = [["단계", "질문", "판정", "도구", "증적(측정/임계)"]]
            for x in sorted_items:
                tool_disp = {"keycloak": "Keycloak", "wazuh": "Wazuh", "nmap": "Nmap",
                             "trivy": "Trivy", "web_probe": "web_probe",
                             "supabase": "Supabase", "vercel": "Vercel", "railway": "Railway",
                             "수동": "수동"}.get(
                    (x.get("tool") or "").lower(), x.get("tool") or "—")
                evidence_str = "—"
                if x.get("metric_value") is not None and x.get("threshold") is not None:
                    evidence_str = f"{x['metric_value']} / {x['threshold']}"
                elif x.get("evidence_summary"):
                    evidence_str = x["evidence_summary"][:80]
                cat_rows.append([
                    Paragraph(x.get("maturity", "-"), CELL),
                    Paragraph((x.get("question") or x.get("item_name") or "-")[:200], CELL),
                    Paragraph(x.get("result", "-"), CELL),
                    Paragraph(tool_disp, CELL),
                    Paragraph(evidence_str, CELL),
                ])
            story.append(_table(cat_rows, F,
                col_widths=[1.6 * cm, CONTENT_W - 9.8 * cm, 1.8 * cm,
                            2.4 * cm, 4.0 * cm],
                header=True, result_col=2, level_col=0,
                small=True))
            story.append(Spacer(1, 0.2 * cm))

        story.append(Spacer(1, 0.15 * cm))

    story.append(Spacer(1, 0.15 * cm))
    story.append(Paragraph(
        "※ 판정 기준 예: 「역할 부여 비율 ≥95% → 충족 / 80~95% → 부분충족 / 미만 → 미충족」. "
        "전체 항목·판정 근거는 10장에 수록.", SMALL))
    story.append(PageBreak())

    # =====================================================================
    # 10. 체크리스트 세부 항목 (펼침)
    # =====================================================================
    story.append(_section_heading_flowable("10", "체크리스트 세부 항목 (펼침)", F, CONTENT_W))
    story.append(Paragraph(
        "대시보드의 3단계 drill-down(필러 → 항목 → 증적)을 문서에서는 펼쳐진 상태로 "
        "전부 수록합니다. 미충족·부분충족·평가불가 항목 우선, 충족 항목은 요약합니다.", BODY))
    story.append(Spacer(1, 0.2 * cm))

    # 미충족 → 부분충족 → 평가불가 → 충족 순으로 정렬
    PRIO = {"미충족": 0, "부분충족": 1, "평가불가": 2, "충족": 3}
    sorted_cr = sorted(cr, key=lambda x: (PRIO.get(x["result"], 9),
                                           x["pillar"], x["item_id"]))

    # 상세 카드: 항목당 표 1개
    for x in sorted_cr:
        story.append(_detail_item_card(x, F, CONTENT_W,
                                        colors, sty, CELL, CELL_B, SMALL, TINY))
        story.append(Spacer(1, 0.12 * cm))

    story.append(PageBreak())

    # =====================================================================
    # 11. 미충족·부분충족·평가불가 집중 정리
    # =====================================================================
    story.append(_section_heading_flowable("11", "미충족·부분충족·평가불가 집중 정리", F, CONTENT_W))
    story.append(Paragraph(
        "점수의 한계와 개선 여지를 한곳에 모아 감사·우선순위 판단을 돕습니다.", BODY))
    story.append(Spacer(1, 0.2 * cm))

    # 11.1
    story.append(Paragraph("11.1 미충족 항목 일람 (필러별 대표)", H_SUB))
    miss_rows = [["필러", "미충족 수", "대표 미충족 항목"]]
    for p in ps:
        misses = [x for x in by_pillar.get(p["pillar"], []) if x["result"] == "미충족"]
        rep = ", ".join(_short_q(m) for m in misses[:3]) if misses else "(미충족 없음)"
        miss_rows.append([_ko_pillar_short(p["pillar"]), str(len(misses)), rep])
    story.append(_table(miss_rows, F,
        col_widths=[3.4 * cm, 2.0 * cm, CONTENT_W - 5.4 * cm],
        header=True))
    story.append(Spacer(1, 0.3 * cm))

    # 11.2
    story.append(Paragraph("11.2 부분충족 — 보완 여지", H_SUB))
    story.append(Paragraph(
        f"총 <b>{partial_cnt}개</b> 부분충족 항목은 \"기능은 존재하나 "
        f"범위·자동화·연동이 미흡\"한 상태로, 비교적 적은 노력으로 충족 전환이 가능합니다.",
        BODY))
    story.append(Spacer(1, 0.3 * cm))

    # 11.3
    story.append(Paragraph("11.3 평가불가 — 사유 및 점수 영향", H_SUB))
    na_reasons = _na_reason_breakdown(cr)
    na_rows = [["평가불가 건수", str(na_cnt)]]
    if na_reasons:
        na_rows.append(["사유 분류 (상위)",
                        " · ".join(f"{k}: {v}건" for k, v in na_reasons.items())])
    else:
        na_rows.append(["사유 분류",
                        ("해당 없음 (데모 모드, 모든 항목 결정론적 평가)"
                         if scan_mode == "demo" and na_cnt == 0
                         else "분류 정보 없음")])
    na_rows.append(["점수 영향",
                    "분모 제외 항목 없음 → 신뢰도 100%" if na_cnt == 0
                    else f"분모에서 {na_cnt}건 제외 → 신뢰도 {confidence_pct}%"])
    story.append(_kv_table(na_rows, F, key_w=3.4 * cm, total_w=CONTENT_W))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(
        "실 스캔 모드에서 도구 미연결·API 오류가 발생하면 해당 항목은 평가불가로 분류되어 "
        "분모에서 제외되며, 신뢰도(confidence)가 100% 미만으로 표시됩니다.", SMALL))

    _set_part("6부 · 개선")
    story.append(PageBreak())

    # =====================================================================
    # 6부 · 개선 — 12. 개선 로드맵
    # =====================================================================
    story.append(_part_badge_flowable("6부 · 개선", F))
    story.append(Spacer(1, 0.4 * cm))

    story.append(_section_heading_flowable("12", "개선 로드맵 (30 / 60 / 90일)", F, CONTENT_W))
    story.append(Paragraph(
        "위험-노력 매트릭스로 우선순위를 정하고, 기간별 과제와 완료 증거를 제시합니다.",
        BODY))
    story.append(Spacer(1, 0.25 * cm))

    # 12.1 위험-노력 매트릭스
    story.append(Paragraph("12.1 위험-노력 매트릭스", H_SUB))
    quadrants = _risk_effort_quadrants(imps)
    mat_rows = [
        [Paragraph("<b><font color='#b91c1c'>Quick Win (고위험·저노력)</font></b><br/>"
                   "<font size='7' color='#6b7280'>우선 착수</font>", CELL),
         Paragraph("<b><font color='#b91c1c'>Major Project (고위험·고노력)</font></b><br/>"
                   "<font size='7' color='#6b7280'>계획 수립</font>", CELL)],
        [Paragraph(quadrants["quick_win"] or "(해당 없음)", CELL),
         Paragraph(quadrants["major"] or "(해당 없음)", CELL)],
        [Paragraph("<b><font color='#475569'>Fill-In (저위험·저노력)</font></b><br/>"
                   "<font size='7' color='#6b7280'>여력 시 처리</font>", CELL),
         Paragraph("<b><font color='#475569'>Thankless (저위험·고노력)</font></b><br/>"
                   "<font size='7' color='#6b7280'>후순위</font>", CELL)],
        [Paragraph(quadrants["fill_in"] or "(해당 없음)", CELL),
         Paragraph(quadrants["thankless"] or "(해당 없음)", CELL)],
    ]
    story.append(Table(mat_rows, colWidths=[CONTENT_W / 2, CONTENT_W / 2],
        style=TableStyle([
            ("FONTNAME",      (0, 0), (-1, -1), F),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#fee2e2")),
            ("BACKGROUND",    (0, 2), (-1, 2), colors.HexColor("#f1f5f9")),
            ("BOX",           (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("INNERGRID",     (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ])))
    story.append(Spacer(1, 0.4 * cm))

    # 12.2 기간별 실행 계획
    story.append(Paragraph("12.2 기간별 실행 계획", H_SUB))
    term_count = {"단기": 0, "중기": 0, "장기": 0}
    for g in imps:
        term_count[g.get("term", "단기")] = term_count.get(g.get("term", "단기"), 0) + 1
    plan_rows = [
        ["기간", "과제 수", "대표 권장 활동", "완료 증거"],
        ["30일 (Quick Win)", str(term_count.get("단기", 0)),
         "평가 범위 확정 · 관리자 MFA 강제 · 권한 목록 정리 · CORS/보안 헤더 점검",
         "권한 표, 설정 캡처, deployment id, 보안 헤더 diff"],
        ["60일 (정착)", str(term_count.get("중기", 0)),
         "API 인증/인가 점검 · 공유 권한 최소화 · secret rotation · Trivy 스캔 정례화",
         "테스트 결과, rotation 로그, Trivy/SCA 리포트"],
        ["90일 (운영)", str(term_count.get("장기", 0)),
         "audit log·보관 기간 정책 확정 · incident runbook 작성 · 재평가 수행",
         "정책 문서, 로그 샘플, 재평가 점수 비교표"],
    ]
    story.append(_table(plan_rows, F,
        col_widths=[3.2 * cm, 1.6 * cm, (CONTENT_W - 4.8 * cm) * 0.55,
                    (CONTENT_W - 4.8 * cm) * 0.45],
        header=True))
    story.append(Spacer(1, 0.4 * cm))

    # 12.3 우선 과제 상세
    story.append(Paragraph("12.3 우선 과제 상세 (상위 3건)", H_SUB))
    top3 = _top_priority(imps, 3)
    if not top3:
        story.append(Paragraph("(우선 과제 데이터 없음)", SMALL))
    else:
        prio_rows = [["위험도", "필러·항목", "도구", "실행 단계 / 기대 효과"]]
        for t in top3:
            tool_disp = {"keycloak": "Keycloak", "wazuh": "Wazuh", "nmap": "Nmap",
                         "trivy": "Trivy", "web_probe": "web_probe",
                         "supabase": "Supabase", "vercel": "Vercel", "railway": "Railway",
                         "수동": "수동"}.get(
                (t.get("tool") or "").lower(), t.get("tool") or "—")
            steps_first = " → ".join(t.get("steps", [])[:3]) or t.get("solution", "")
            effect = t.get("expected_effect") or t.get("expected_gain") or ""
            cell_html = f"{steps_first}"
            if effect:
                cell_html += f"<br/><font color='#6b7280' size='7'>기대 효과: {effect}</font>"
            prio_rows.append([
                t.get("priority", "Medium"),
                f"{_ko_pillar_short(t.get('pillar', ''))} · {t.get('item_id') or t.get('category', '')}",
                tool_disp,
                Paragraph(cell_html, CELL),
            ])
        story.append(_table(prio_rows, F,
            col_widths=[1.8 * cm, 5.0 * cm, 2.0 * cm, CONTENT_W - 8.8 * cm],
            header=True, priority_col=0))

    _set_part("7부 · 부록")
    story.append(PageBreak())

    # =====================================================================
    # 7부 · 부록
    # =====================================================================
    story.append(_part_badge_flowable("7부 · 부록", F))
    story.append(Spacer(1, 0.4 * cm))

    # 13. 표준 매핑
    story.append(_section_heading_flowable("13", "표준 매핑", F, CONTENT_W))
    story.append(Paragraph(
        "진단 항목을 국제 표준에 대응시켜 컴플라이언스 대응에 활용합니다. "
        "(내보내기: <b>JSON / CSV</b> — API <font face='Courier'>GET /api/report/standards/{session_id}</font>)",
        BODY))
    story.append(Spacer(1, 0.2 * cm))
    std_table_rows = [
        ["필러", "NIST SP 800-207 매핑 원칙", "CIS Controls v8 매핑"],
    ]
    pillar_to_nist_label = {
        "식별자 및 신원":         "Tenets #3·#4·#6 (ICAM·dynamic policy·auth)",
        "기기 및 엔드포인트":     "Tenets #1·#5·#7 (resource·integrity·monitoring)",
        "네트워크":               "Tenets #2·#7 (secured communication·monitoring)",
        "시스템":                 "Tenets #4·#5·#6 (policy·integrity·auth)",
        "애플리케이션 및 워크로드": "Tenets #1·#4·#6",
        "데이터":                 "Tenets #1·#3 (resource·per-session)",
    }
    pillar_to_cis_label = {
        "식별자 및 신원":         "CIS 5(Account), 6(Access Control)",
        "기기 및 엔드포인트":     "CIS 1(Assets), 2(SW), 4(Secure Config)",
        "네트워크":               "CIS 12(Network), 13(Network Monitoring)",
        "시스템":                 "CIS 4(Config), 8(Audit Log), 17(Incident)",
        "애플리케이션 및 워크로드": "CIS 2(SW), 7(Vuln Mgmt), 16(App Security)",
        "데이터":                 "CIS 3(Data Protection), 11(Data Recovery)",
    }
    for p in ps:
        std_table_rows.append([
            _ko_pillar_short(p["pillar"]),
            pillar_to_nist_label.get(p["pillar"], "—"),
            pillar_to_cis_label.get(p["pillar"], "—"),
        ])
    story.append(_table(std_table_rows, F,
        col_widths=[3.2 * cm, (CONTENT_W - 3.2 * cm) * 0.5,
                    (CONTENT_W - 3.2 * cm) * 0.5],
        header=True, small=True))
    story.append(Spacer(1, 0.4 * cm))

    # 14. 증적 레지스터
    story.append(_section_heading_flowable("14", "증적 레지스터", F, CONTENT_W))
    story.append(Paragraph(
        "각 판정의 근거 자료 목록입니다. (xlsx 내보내기: API "
        "<font face='Courier'>GET /api/report/evidence-register/{session_id}</font>)",
        BODY))
    story.append(Spacer(1, 0.2 * cm))

    ev_rows = [["항목 ID", "증적 유형", "출처 / 식별자", "수집 시각"]]
    # 상위 15건만 (지면 절약)
    ev_items = [x for x in cr if x.get("evidence_summary") or x.get("metric_key")][:15]
    if not ev_items:
        ev_rows.append(["—", "—", "(수집된 증적 없음)", "—"])
    else:
        for x in ev_items:
            tool_disp = {"keycloak": "Keycloak", "wazuh": "Wazuh", "nmap": "Nmap",
                         "trivy": "Trivy", "web_probe": "web_probe",
                         "supabase": "Supabase", "vercel": "Vercel", "railway": "Railway",
                         "수동": "수동"}.get(
                (x.get("tool") or "").lower(), x.get("tool") or "—")
            ev_type = "API 응답" if tool_disp in ("Keycloak", "Wazuh", "Supabase") else (
                "스캔 결과" if tool_disp in ("Nmap", "Trivy", "web_probe") else
                "플랫폼 API" if tool_disp in ("Vercel", "Railway") else "수동 입력")
            src = f"{tool_disp} · {x.get('metric_key', '') or '(metric 없음)'}"
            cat = (x.get("collected_at") or "")[:10]
            ev_rows.append([x["item_id"], ev_type, src[:90], cat or "—"])
    story.append(_table(ev_rows, F,
        col_widths=[2.6 * cm, 2.4 * cm, CONTENT_W - 7.6 * cm, 2.6 * cm],
        header=True, small=True))
    story.append(Spacer(1, 0.3 * cm))

    # 15. 판정 로그
    story.append(_section_heading_flowable("15", "판정 로그", F, CONTENT_W))
    story.append(Paragraph(
        "부분충족·평가불가 항목의 판정 사유입니다. "
        "(Markdown 내보내기: API <font face='Courier'>GET /api/report/decision-log/{session_id}</font>)",
        BODY))
    story.append(Spacer(1, 0.2 * cm))
    debate = [x for x in cr if x["result"] in ("부분충족", "평가불가")][:15]
    dl_rows = [["항목 ID", "판정", "판정 사유"]]
    if not debate:
        dl_rows.append(["—", "—", "(부분충족·평가불가 항목 없음)"])
    else:
        for x in debate:
            reason = x.get("reason_label") or x.get("evidence_summary") \
                or x.get("auto_error") or "(사유 데이터 없음)"
            dl_rows.append([x["item_id"], x["result"], reason[:140]])
    story.append(_table(dl_rows, F,
        col_widths=[2.6 * cm, 2.0 * cm, CONTENT_W - 4.6 * cm],
        header=True, result_col=1, small=True))
    story.append(Spacer(1, 0.4 * cm))

    # 16. OCSF
    story.append(_section_heading_flowable("16", "OCSF 내보내기 안내", F, CONTENT_W))
    story.append(Paragraph(
        "규제 대응을 위해 진단 결과를 OCSF(Open Cybersecurity Schema Framework) 형식 "
        "JSON으로 내보낼 수 있습니다. 대시보드 또는 API "
        "<font face='Courier'>GET /api/assessment/ocsf/{session_id}</font>", BODY))
    story.append(Spacer(1, 0.4 * cm))

    # 17. 용어집
    story.append(_section_heading_flowable("17", "용어집", F, CONTENT_W))
    glossary = [
        ["용어", "정의"],
        ["성숙도 4단계",         "기존(1.5↓) / 초기(1.5~2.5) / 향상(2.5~3.5) / 최적화(3.5↑)"],
        ["신뢰도(confidence)",   "평가 가능 항목 ÷ 전체 항목 × 100%. 도구 미연결 시 하락."],
        ["평가불가",             "도구 미연결·오류로 판정 불가한 항목. 점수 분모에서 제외."],
        ["위험-노력 매트릭스",   "위험도·투입 노력 기준 4사분면 분류(Quick Win / Major / Fill-In / Thankless)."],
        ["SCA",                  "Security Configuration Assessment — Wazuh의 보안 설정 준수 평가."],
        ["FIM",                  "File Integrity Monitoring — 파일 무결성 모니터링."],
        ["ICAM",                 "Identity, Credential & Access Management — 신원·자격·접근 통합 관리."],
        ["RBAC / ABAC",          "역할/속성 기반 접근 통제 모델."],
        ["MFA / WebAuthn",       "다중 인증 / FIDO 표준 기반 무비밀번호 인증."],
        ["SBOM",                 "Software Bill of Materials — 소프트웨어 구성 명세."],
    ]
    story.append(_table(glossary, F,
        col_widths=[4.0 * cm, CONTENT_W - 4.0 * cm], header=True, small=True))
    story.append(Spacer(1, 0.5 * cm))

    story.append(HRFlowable(width=CONTENT_W, thickness=0.5,
                            color=colors.HexColor("#cbd5e1")))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(
        "— 본 보고서 끝 — &nbsp;&nbsp;|&nbsp;&nbsp; "
        "Readyz-T Zero Trust Assessment &nbsp;|&nbsp; 신뢰많이된다 팀",
        sty("end_footer", fontSize=8.5, leading=11, alignment=1,
            textColor=colors.HexColor("#64748b"))))

    doc.build(story)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Helper Flowables / 테이블 유틸 (PDF용)
# ─────────────────────────────────────────────────────────────────────────────

from reportlab.platypus import Flowable, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors as _rl_colors
from reportlab.lib.styles import ParagraphStyle


class _PartMarker(Flowable):
    """현재 부 이름을 헤더 함수에 전달하기 위한 invisible flowable.

    story 진행 중 페이지가 넘어갈 때 BaseDocTemplate 의 onPage 콜백이 호출되는데,
    이때 closest preceding _PartMarker.name 값을 current_part 에 반영하는 방식.
    Flowable.draw() 시점에 부 이름을 갱신.
    """
    def __init__(self, name: str, target_dict: dict):
        super().__init__()
        self.name = name
        self.target = target_dict

    def wrap(self, *args):
        return (0, 0)

    def draw(self):
        self.target["name"] = self.name


def _table(rows, font, col_widths, header=False, result_col=None,
           level_col=None, priority_col=None, sum_row=False, small=False):
    """표준 표 스타일.

    rows[0]가 헤더(header=True), result_col / level_col / priority_col 위치에
    색상 강조. sum_row=True 이면 마지막 행을 합계로 처리.
    """
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    size = 7.5 if small else 8.2
    ts = [
        ("FONTNAME",     (0, 0), (-1, -1), font),
        ("FONTSIZE",     (0, 0), (-1, -1), size),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("BOX",          (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("INNERGRID",    (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f8fafc")]),
    ]
    if header:
        ts += [
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), font),
            ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
        ]
    # 결과 컬럼 색
    result_palette = {
        "충족":    ("#dcfce7", "#15803d"),
        "부분충족": ("#fef9c3", "#92400e"),
        "미충족":  ("#fee2e2", "#b91c1c"),
        "평가불가": ("#f1f5f9", "#475569"),
    }
    level_palette = {
        "기존":   ("#fee2e2", "#b91c1c"),
        "초기":   ("#fef9c3", "#92400e"),
        "향상":   ("#dbeafe", "#1e40af"),
        "최적화": ("#dcfce7", "#15803d"),
    }
    priority_palette = {
        "Critical": ("#fee2e2", "#b91c1c"),
        "High":     ("#fff7ed", "#c2410c"),
        "Medium":   ("#fef9c3", "#92400e"),
        "Low":      ("#f1f5f9", "#475569"),
    }

    def _apply_color(col_idx, palette):
        if col_idx is None:
            return
        for ri in range(1, len(rows)):
            raw = rows[ri][col_idx]
            label = raw if isinstance(raw, str) else (
                raw.text if hasattr(raw, "text") else str(raw))
            if label in palette:
                bg, fg = palette[label]
                ts.append(("BACKGROUND", (col_idx, ri), (col_idx, ri),
                           colors.HexColor(bg)))
                ts.append(("TEXTCOLOR",  (col_idx, ri), (col_idx, ri),
                           colors.HexColor(fg)))

    _apply_color(result_col, result_palette)
    _apply_color(level_col, level_palette)
    _apply_color(priority_col, priority_palette)

    if sum_row and len(rows) > 1:
        last = len(rows) - 1
        ts.append(("BACKGROUND", (0, last), (-1, last), colors.HexColor("#e2e8f0")))
        ts.append(("FONTNAME",   (0, last), (-1, last), font))

    return Table(rows, colWidths=col_widths, style=TableStyle(ts), repeatRows=1 if header else 0)


def _kv_table(rows, font, key_w, total_w):
    """좌측 키(회색)·우측 값 정렬 2-컬럼 표."""
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    return Table(rows, colWidths=[key_w, total_w - key_w],
        style=TableStyle([
            ("FONTNAME",     (0, 0), (-1, -1), font),
            ("FONTSIZE",     (0, 0), (-1, -1), 8.5),
            ("VALIGN",       (0, 0), (-1, -1), "TOP"),
            ("TEXTCOLOR",    (0, 0), (0, -1), colors.HexColor("#6b7280")),
            ("TEXTCOLOR",    (1, 0), (1, -1), colors.HexColor("#111827")),
            ("BACKGROUND",   (0, 0), (0, -1), colors.HexColor("#f8fafc")),
            ("LINEBELOW",    (0, 0), (-1, -2), 0.3, colors.HexColor("#e5e7eb")),
            ("BOX",          (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
            ("LEFTPADDING",  (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING",   (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ]))


def _callout_box(html: str, body_style, font, total_w,
                 bg="#f8fafc", border="#cbd5e1"):
    """배경색·테두리 박스 안에 Paragraph."""
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib import colors
    p = Paragraph(html, body_style)
    return Table([[p]], colWidths=[total_w],
        style=TableStyle([
            ("FONTNAME",     (0, 0), (-1, -1), font),
            ("BACKGROUND",   (0, 0), (-1, -1), colors.HexColor(bg)),
            ("BOX",          (0, 0), (-1, -1), 0.5, colors.HexColor(border)),
            ("LEFTPADDING",  (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING",   (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 8),
        ]))


def _stat_cell(value, label, fg, bg, font):
    """3.1 핵심 지표 카드 1개 — 흰 배경 + 색 숫자 + 회색 라벨 (샘플 매칭).
    bg 인자는 backwards-compat 으로 받되 사용하지 않음 (샘플은 흰 배경).
    """
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    big = ParagraphStyle("big", fontName=font, fontSize=24, leading=28, alignment=1,
                         textColor=colors.HexColor(fg))
    lab = ParagraphStyle("lab", fontName=font, fontSize=9, leading=12, alignment=1,
                         textColor=colors.HexColor("#64748b"))
    # <font color> 명시 — 한글 폰트에 bold variant 없을 때 textColor가 누락되는 경우 방지.
    return Table(
        [[Paragraph(f'<font color="{fg}">{value}</font>', big)],
         [Paragraph(label, lab)]],
        style=TableStyle([
            ("BACKGROUND",   (0, 0), (-1, -1), colors.white),
            ("BOX",          (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
            ("LINEBELOW",    (0, 0), (-1, 0), 0, colors.white),
            ("ROUNDEDCORNERS",(2, 2, 2, 2)),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING",   (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 10),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ]),
    )


def _detail_item_card(item: dict, font, total_w, colors_mod, sty_fn,
                      CELL, CELL_B, SMALL, TINY):
    """10장 체크리스트 세부 항목 카드 1개."""
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib import colors

    result = item.get("result", "-")
    result_palette = {
        "충족":    ("#dcfce7", "#15803d"),
        "부분충족": ("#fef9c3", "#92400e"),
        "미충족":  ("#fee2e2", "#b91c1c"),
        "평가불가": ("#f1f5f9", "#475569"),
    }
    bg, fg = result_palette.get(result, ("#ffffff", "#374151"))

    tool_disp = {"keycloak": "Keycloak", "wazuh": "Wazuh", "nmap": "Nmap",
                 "trivy": "Trivy", "web_probe": "web_probe",
                 "supabase": "Supabase", "vercel": "Vercel", "railway": "Railway",
                 "수동": "수동"}.get(
        (item.get("tool") or "").lower(), item.get("tool") or "—")

    header_html = (
        f"<b><font color='#1e3a5f'>{item.get('item_id','-')}</font></b>  "
        f"<font color='#6b7280' size='7'>{item.get('category','')} · 단계 "
        f"<b>{item.get('maturity','-')}</b></font>"
        f"&nbsp;&nbsp;&nbsp;"
        f"<font size='7' color='#6b7280'>도구 <b>{tool_disp}</b></font>"
    )
    result_html = (
        f"<font color='{fg}'><b>{result}</b></font>  "
        f"<font color='#6b7280' size='7'>(성숙도 점수 {item.get('score', 0.0):.2f} / "
        f"가중치 {(item.get('maturity_score') or 0) / 10:.1f})</font>"
    )

    metric_html = ""
    mv = item.get("metric_value")
    th = item.get("threshold")
    if mv is not None and th is not None:
        ratio = (mv / th) if th else 0.0
        ratio_pct = max(0, min(100, int(ratio * 100)))
        metric_html = (
            f"<font color='#374151'>{item.get('metric_key','metric')} = "
            f"<b>{mv}</b> / 임계값 <b>{th}</b></font>  "
            f"<font color='#6b7280' size='7'>· 진척 {ratio_pct}%</font>"
        )
    elif item.get("evidence_summary"):
        metric_html = item.get("evidence_summary")[:240]
    else:
        metric_html = "<font color='#9ca3af'>(수집된 증적 없음)</font>"

    rows = [
        [Paragraph(header_html, CELL)],
        [Paragraph(f"<b>질문</b>  {item.get('question') or item.get('item_name') or '-'}", CELL_B)],
        [Paragraph(f"<b>판정 결과</b>  {result_html}", CELL)],
        [Paragraph(f"<b>판정 기준</b>  <font color='#6b7280'>{item.get('criteria') or '-'}</font>", CELL)],
        [Paragraph(f"<b>증적 · 근거</b>  {metric_html}", CELL)],
    ]
    if item.get("raw_json_summary"):
        rows.append([Paragraph(
            f"<b>raw 응답 요약</b>  <font face='Courier' size='7' color='#475569'>"
            f"{item['raw_json_summary']}</font>", CELL)])
    if item.get("auto_error"):
        rows.append([Paragraph(
            f"<font color='#b91c1c'><b>오류</b></font>  "
            f"<font color='#b91c1c' size='7'>{item['auto_error'][:200]}</font>", CELL)])
    if item.get("evidence_observed"):
        rows.append([Paragraph(
            f"<b>수동 관찰</b>  {item['evidence_observed'][:240]}", CELL)])

    return Table(rows, colWidths=[total_w], style=TableStyle([
        ("FONTNAME",     (0, 0), (-1, -1), font),
        ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor(bg)),
        ("BOX",          (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("LINEBELOW",    (0, 0), (-1, 0), 0.5, colors.HexColor(fg)),
        ("LEFTPADDING",  (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
    ]))


def _weak_pillar_reasons(weak_pillars: list[dict], cr: list[dict]) -> dict:
    """가장 취약한 필러 1~3개의 대표 미충족 사유 요약."""
    by_pillar: dict[str, list] = {}
    for r in cr:
        if r["result"] in ("미충족", "부분충족"):
            by_pillar.setdefault(r["pillar"], []).append(r)
    out = {}
    for p in weak_pillars:
        items = by_pillar.get(p["pillar"], [])[:3]
        if not items:
            out[p["pillar"]] = "(취약 항목 없음)"
        else:
            short = ", ".join(_short_q(x) for x in items)
            out[p["pillar"]] = short
    return out


def _short_q(item: dict, max_len: int = 50) -> str:
    q = item.get("question") or item.get("item_name") or item.get("item_id") or "-"
    if len(q) > max_len:
        q = q[:max_len - 1] + "…"
    return q


def _urgent_item(items: list[dict]) -> str:
    misses = [x for x in items if x["result"] == "미충족"]
    if not misses:
        partials = [x for x in items if x["result"] == "부분충족"]
        if not partials:
            return "(시급 항목 없음 — 필러 점수 양호)"
        return ", ".join(_short_q(x, 40) for x in partials[:3])
    return ", ".join(_short_q(x, 40) for x in misses[:3])


def _quick_wins(imps: list[dict], n: int) -> list[dict]:
    """단기(30일) + Critical/High 우선."""
    PRIO = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    short_term = [g for g in imps if g.get("term") == "단기"]
    short_term.sort(key=lambda g: (PRIO.get(g.get("priority"), 9),
                                    g.get("pillar") or ""))
    if len(short_term) >= n:
        return short_term[:n]
    # 부족하면 전체에서 보충
    rest = sorted(imps, key=lambda g: (PRIO.get(g.get("priority"), 9),
                                        g.get("pillar") or ""))
    seen = {id(x) for x in short_term}
    for x in rest:
        if id(x) in seen:
            continue
        short_term.append(x)
        if len(short_term) >= n:
            break
    return short_term[:n]


def _tool_auto_counts(cr: list[dict]) -> dict:
    out: dict = {}
    for x in cr:
        t = (x.get("tool") or "").lower() or "수동"
        if t == "manual":
            t = "수동"
        out[t] = out.get(t, 0) + 1
    return out


def _next_level_gap(score: float) -> tuple[Optional[str], float]:
    if score < 1.5:
        return ("초기", 1.5 - score)
    if score < 2.5:
        return ("향상", 2.5 - score)
    if score < 3.5:
        return ("최적화", 3.5 - score)
    return (None, 0.0)


def _risk_effort_quadrants(imps: list[dict]) -> dict:
    """위험-노력 매트릭스 4분면 — 각 칸에 항목 리스트(간단 텍스트)."""
    PRIO = {"Critical": 3, "High": 2, "Medium": 1, "Low": 0}
    out = {"quick_win": [], "major": [], "fill_in": [], "thankless": []}
    for g in imps:
        risk = PRIO.get(g.get("priority"), 0)
        # 노력 추정: 단기=저, 중기=중, 장기=고
        eff_short = g.get("term") == "단기"
        eff_long = g.get("term") == "장기"
        task = g.get("task") or "(과제명 없음)"
        task_short = task[:60] + ("…" if len(task) > 60 else "")
        bucket = ("quick_win" if risk >= 2 and eff_short else
                  "major" if risk >= 2 and eff_long else
                  "fill_in" if risk < 2 and eff_short else
                  "thankless")
        out[bucket].append(f"• {task_short}")
    return {k: "<br/>".join(v[:5]) for k, v in out.items()}


def _top_priority(imps: list[dict], n: int) -> list[dict]:
    PRIO = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    return sorted(imps, key=lambda g: (PRIO.get(g.get("priority"), 9),
                                        g.get("term") == "장기"))[:n]


def _na_reason_breakdown(cr: list[dict]) -> dict:
    counts: dict = {}
    for x in cr:
        if x["result"] != "평가불가":
            continue
        label = x.get("reason_label") or x.get("auto_error") or "(사유 미기재)"
        # 상위 키워드만
        key = label[:40]
        counts[key] = counts.get(key, 0) + 1
    # top 5
    return dict(sorted(counts.items(), key=lambda kv: -kv[1])[:5])


# ── 라우터 ──────────────────────────────────────────────────────────────────

@router.get("/generate")
async def generate_report_by_query(
    session_id: int,
    fmt: str = Query(default="json", pattern="^(json|pdf)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    session = db.query(DiagnosisSession).filter(DiagnosisSession.session_id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
    assert_session_access(current_user, session)

    data = _build_data(session_id, db)
    if fmt == "pdf":
        pdf_bytes = await asyncio.to_thread(_make_pdf, data)
        filename = f"zt-report-{session_id}-{data['generated_at'][:10]}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    return JSONResponse(content={
        "report_generated_at": data["generated_at"],
        "session": data["session"],
        "summary": data["summary"],
        "pillar_scores": data["pillar_scores"],
        "checklist_results": data["checklist_results"],
        "improvement_targets": data["improvement_targets"],
    })


@router.get("/generate/{session_id}")
async def generate_report(
    session_id: int,
    fmt: str = Query(default="json", pattern="^(json|pdf)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return await generate_report_by_query(
        session_id=session_id, fmt=fmt, db=db, current_user=current_user,
    )


@router.get("/standards/{session_id}")
def get_standards_mapping(
    session_id: int,
    fmt: str = Query(default="json", pattern="^(json|csv)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """진단 세션의 NIST 800-207 / CIS Controls v8 매핑 export.

    fmt=json → 표준별 충족·미충족 집계 + 세부 매핑
    fmt=csv  → 표준별 1행씩, columns: standard, id, title, pass, fail, na, compliance_rate
    """
    session = db.query(DiagnosisSession).filter(
        DiagnosisSession.session_id == session_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
    assert_session_access(current_user, session)

    rows = (
        db.query(DiagnosisResult, Checklist)
        .join(Checklist, DiagnosisResult.check_id == Checklist.check_id)
        .filter(DiagnosisResult.session_id == session_id)
        .all()
    )
    checklist_results = [
        {
            "item_id": cl.item_id, "pillar": cl.pillar, "category": cl.category,
            "item": cl.item_name, "result": dr.result,
        }
        for dr, cl in rows
    ]

    summary = session_standards_summary(checklist_results)

    if fmt == "csv":
        import csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["standard", "id", "title", "pass", "fail", "na", "compliance_rate"])
        for n in summary["nist_800_207"]:
            w.writerow(["NIST 800-207", n["tenet"], n["title"], n["pass"], n["fail"], n["na"], n["compliance_rate"]])
        for c in summary["cis_controls_v8"]:
            w.writerow(["CIS Controls v8", c["control_id"], c["title"], c["pass"], c["fail"], c["na"], c["compliance_rate"]])
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="zt-standards-{session_id}.csv"'},
        )

    return JSONResponse(content={
        "session_id": session_id,
        "nist_800_207": summary["nist_800_207"],
        "cis_controls_v8": summary["cis_controls_v8"],
        "generated_at": datetime.now(timezone.utc).isoformat(),
    })


# ─── 증적 목록 xlsx (가이드 §7 산출물) ──────────────────────────────────────

def _build_evidence_register_xlsx(session_id: int, db: Session) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    session = db.query(DiagnosisSession).filter(
        DiagnosisSession.session_id == session_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
    org = db.query(Organization).filter(Organization.org_id == session.org_id).first()
    user = db.query(User).filter(User.user_id == session.user_id).first()

    results = (
        db.query(DiagnosisResult, Checklist)
        .join(Checklist, DiagnosisResult.check_id == Checklist.check_id)
        .filter(DiagnosisResult.session_id == session_id)
        .all()
    )
    by_check: dict[int, dict] = {}
    for dr, cl in results:
        by_check[cl.check_id] = {"dr": dr, "cl": cl}

    evidences = db.query(Evidence).filter(Evidence.session_id == session_id).all()
    collected = db.query(CollectedData).filter(CollectedData.session_id == session_id).all()
    coll_by_check: dict[int, CollectedData] = {c.check_id: c for c in collected}

    ev_by_check: dict[int, list[Evidence]] = {}
    for ev in evidences:
        ev_by_check.setdefault(ev.check_id, []).append(ev)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "증적 목록"

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF1E3A5F")
    sub_font    = Font(name="Arial", size=9)
    sub_fill_a  = PatternFill("solid", fgColor="FFFFFFFF")
    sub_fill_b  = PatternFill("solid", fgColor="FFF8FAFC")
    result_fill = {
        "충족":      PatternFill("solid", fgColor="FFDCFCE7"),
        "부분충족":  PatternFill("solid", fgColor="FFFEF9C3"),
        "미충족":    PatternFill("solid", fgColor="FFFEE2E2"),
        "평가불가":  PatternFill("solid", fgColor="FFF3F4F6"),
    }
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    headers = [
        "항목ID", "Pillar", "카테고리", "항목명", "성숙도",
        "진단결과", "점수", "출처(자동/수동)", "도구",
        "증적유형", "관찰 내용", "파일명", "파일크기(B)", "MIME",
        "위치/URL", "원인/근거", "영향도", "수집·등록 시각",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(1, col, h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    widths = [12, 12, 22, 36, 8, 10, 6, 14, 10, 14, 50, 28, 12, 16, 36, 36, 8, 22]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    row = 2
    for check_id in sorted(by_check.keys()):
        entry = by_check[check_id]
        cl = entry["cl"]
        dr = entry["dr"]
        evs = ev_by_check.get(check_id, [])
        coll = coll_by_check.get(check_id)
        tool = coll.tool if coll else ""
        source = "수동" if tool == "수동" else ("자동" if tool else "")
        coll_at = coll.collected_at.isoformat() if coll and coll.collected_at else ""

        auto_obs = ""
        if coll and isinstance(coll.raw_json, dict):
            rj = coll.raw_json
            parts = []
            if "score" in rj:
                parts.append(f"score={rj.get('score')}")
            if "metric_value" in rj or coll.metric_value is not None:
                mv = rj.get("metric_value", coll.metric_value)
                parts.append(f"{coll.metric_key}={mv}/{coll.threshold}")
            if rj.get("issues"):
                parts.append(f"issues={len(rj['issues'])}건")
            auto_obs = " · ".join(parts)

        def _write_row(r, ev: Optional[Evidence]):
            vals = [
                cl.item_id, cl.pillar, cl.category, cl.item_name, cl.maturity,
                dr.result, round(dr.score or 0.0, 2),
                source, tool,
                (ev.source if ev else "자동수집") if (ev or coll) else "",
                (ev.observed if ev and ev.observed else "") or auto_obs,
                ev.original_filename if ev and ev.original_filename else "",
                ev.file_size if ev and ev.file_size else "",
                ev.mime_type if ev and ev.mime_type else "",
                ev.location if ev and ev.location else "",
                ev.reason if ev and ev.reason else "",
                ev.impact if ev and ev.impact is not None else "",
                coll_at,
            ]
            fill_zebra = sub_fill_b if (r % 2 == 0) else sub_fill_a
            for col, v in enumerate(vals, start=1):
                c = ws.cell(r, col, v)
                c.font = sub_font
                c.alignment = wrap
                c.fill = fill_zebra
            rcell = ws.cell(r, 6)
            if dr.result in result_fill:
                rcell.fill = result_fill[dr.result]
                rcell.font = Font(name="Arial", size=9, bold=True)

        if evs:
            for ev in evs:
                _write_row(row, ev)
                row += 1
        else:
            _write_row(row, None)
            row += 1

    ws_meta = wb.create_sheet("메타")
    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 60
    meta_rows = [
        ("기관",           org.name if org else ""),
        ("담당자",         user.name if user else ""),
        ("세션 ID",        session.session_id),
        ("진단 시작",      session.started_at.isoformat() if session.started_at else ""),
        ("진단 완료",      session.completed_at.isoformat() if session.completed_at else ""),
        ("상태",           session.status),
        ("총 항목 수",     len(by_check)),
        ("증적 파일 수",   sum(1 for ev in evidences if ev.file_path)),
        ("자동 수집 수",   len([c for c in collected if c.tool != "수동"])),
        ("수동 등록 수",   len([c for c in collected if c.tool == "수동"])),
        ("생성 시각",      datetime.now(timezone.utc).isoformat()),
    ]

    em = build_evaluation_meta(session)
    eval_ver = em.get("evaluation_version") or {}
    scope_as = em.get("evaluation_scope_assets") or []
    data_cls = em.get("data_classifications") or []
    reviewers_m = em.get("reviewers") or {}

    if em.get("scan_mode"):
        meta_rows.append(("진단 모드", em.get("scan_mode")))
    if em.get("selected_tools"):
        meta_rows.append(("수행 도구", ", ".join(em.get("selected_tools"))))
    if em.get("excluded_tools"):
        meta_rows.append(("제외 도구", ", ".join(em.get("excluded_tools"))))
    if eval_ver.get("version_label"):
        meta_rows.append(("버전 라벨", eval_ver["version_label"]))
    if eval_ver.get("git_commit"):
        meta_rows.append(("Git commit", eval_ver["git_commit"]))
    if eval_ver.get("frontend_deployment"):
        meta_rows.append(("Frontend deployment", eval_ver["frontend_deployment"]))
    if eval_ver.get("backend_deployment"):
        meta_rows.append(("Backend deployment", eval_ver["backend_deployment"]))
    if reviewers_m.get("app_owner"):
        meta_rows.append(("App owner", reviewers_m["app_owner"]))
    if reviewers_m.get("backend_owner"):
        meta_rows.append(("Backend owner", reviewers_m["backend_owner"]))
    if reviewers_m.get("cloud_owner"):
        meta_rows.append(("Cloud owner", reviewers_m["cloud_owner"]))
    if reviewers_m.get("security_reviewer"):
        meta_rows.append(("Security reviewer", reviewers_m["security_reviewer"]))

    for r, (k, v) in enumerate(meta_rows, start=1):
        ws_meta.cell(r, 1, k).font = Font(name="Arial", size=10, bold=True, color="FF617087")
        ws_meta.cell(r, 2, str(v) if v is not None else "")

    if scope_as:
        ws_assets = wb.create_sheet("평가 범위 자산")
        ws_assets.column_dimensions["A"].width = 8
        ws_assets.column_dimensions["B"].width = 24
        ws_assets.column_dimensions["C"].width = 60
        ws_assets.cell(1, 1, "포함").font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        ws_assets.cell(1, 2, "자산").font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        ws_assets.cell(1, 3, "값").font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        for c in (ws_assets.cell(1, 1), ws_assets.cell(1, 2), ws_assets.cell(1, 3)):
            c.fill = PatternFill("solid", fgColor="FF1E3A5F")
        for ri, a in enumerate(scope_as, start=2):
            ws_assets.cell(ri, 1, "✓" if a.get("included") else "—")
            ws_assets.cell(ri, 2, a.get("name", ""))
            ws_assets.cell(ri, 3, a.get("value", ""))

    if data_cls:
        ws_data = wb.create_sheet("데이터 등급")
        ws_data.column_dimensions["A"].width = 10
        ws_data.column_dimensions["B"].width = 28
        ws_data.column_dimensions["C"].width = 40
        sens_fill = {"높음": "FFFEE2E2", "중간": "FFFEF9C3", "낮음": "FFF3F4F6"}
        ws_data.cell(1, 1, "민감도").font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        ws_data.cell(1, 2, "데이터 항목").font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        ws_data.cell(1, 3, "보관 위치").font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        for c in (ws_data.cell(1, 1), ws_data.cell(1, 2), ws_data.cell(1, 3)):
            c.fill = PatternFill("solid", fgColor="FF1E3A5F")
        for ri, d in enumerate(data_cls, start=2):
            sens = d.get("sensitivity", "")
            sc = ws_data.cell(ri, 1, sens)
            if sens in sens_fill:
                sc.fill = PatternFill("solid", fgColor=sens_fill[sens])
            ws_data.cell(ri, 2, d.get("name", ""))
            ws_data.cell(ri, 3, d.get("storage_location", ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/evidence-register/{session_id}")
async def download_evidence_register(
    session_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """증적 목록 xlsx 다운로드 — 가이드 §7 산출물 "evidence_register.xlsx".

    각 진단 결과에 연결된 자동 수집(CollectedData) + 수동 등록 증적(Evidence)
    을 한 xlsx 로 정리. 컬럼: 항목/결과/출처/증적유형/파일/위치/관찰내용/시각 등.
    """
    session = db.query(DiagnosisSession).filter(
        DiagnosisSession.session_id == session_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
    assert_session_access(current_user, session)

    try:
        data = await asyncio.to_thread(_build_evidence_register_xlsx, session_id, db)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("[report] evidence register build failed: %s", exc)
        raise HTTPException(status_code=500, detail="증적 목록 생성에 실패했습니다.")

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = f"evidence-register-{session_id}-{today}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── 판정 로그 markdown (가이드 §7 산출물 decision_log.md) ──────────────────

def _build_decision_log_md(session_id: int, db: Session) -> str:
    session = db.query(DiagnosisSession).filter(
        DiagnosisSession.session_id == session_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
    org = db.query(Organization).filter(Organization.org_id == session.org_id).first()
    user = db.query(User).filter(User.user_id == session.user_id).first()

    results = (
        db.query(DiagnosisResult, Checklist)
        .join(Checklist, DiagnosisResult.check_id == Checklist.check_id)
        .filter(DiagnosisResult.session_id == session_id)
        .all()
    )
    evidences = db.query(Evidence).filter(Evidence.session_id == session_id).all()
    collected = db.query(CollectedData).filter(CollectedData.session_id == session_id).all()
    ev_by_check: dict[int, list[Evidence]] = {}
    for ev in evidences:
        ev_by_check.setdefault(ev.check_id, []).append(ev)
    coll_by_check: dict[int, CollectedData] = {c.check_id: c for c in collected}

    eval_meta = build_evaluation_meta(session)

    total = len(results)
    pass_cnt    = sum(1 for dr, _ in results if dr.result == "충족")
    partial_cnt = sum(1 for dr, _ in results if dr.result == "부분충족")
    fail_cnt    = sum(1 for dr, _ in results if dr.result == "미충족")
    na_cnt      = sum(1 for dr, _ in results if dr.result == "평가불가")

    debate_results = [(dr, cl) for dr, cl in results
                      if dr.result in ("부분충족", "평가불가")]
    debate_results.sort(key=lambda x: (x[1].pillar or "", x[1].item_id or ""))

    lines: list[str] = []

    lines.append(f"# 판정 로그 — {org.name if org else '(미상)'}")
    lines.append("")
    lines.append(f"> 산출물 패키지 — `decision_log.md`")
    lines.append(f"> 논쟁 가능 항목(부분충족·평가불가)의 판단 근거와 리뷰어 의견을 기록합니다.")
    lines.append("")
    lines.append("## 1. 기본 정보")
    lines.append("")
    lines.append(f"- **세션 ID**: {session.session_id}")
    lines.append(f"- **담당자**: {user.name if user else '(미상)'}")
    lines.append(f"- **진단 시작**: {session.started_at.isoformat() if session.started_at else '-'}")
    lines.append(f"- **진단 완료**: {session.completed_at.isoformat() if session.completed_at else '-'}")
    lines.append(f"- **상태**: {session.status}")
    lines.append(f"- **생성 시각**: {datetime.now(timezone.utc).isoformat()}")
    lines.append("")

    lines.append("## 2. 평가 범위 / 모드")
    lines.append("")
    lines.append(f"- **진단 모드**: {eval_meta.get('scan_mode', 'demo')}")
    lines.append(f"- **사용 환경**: IdP={eval_meta.get('profile_select', {}).get('idp_type', 'none')} / "
                 f"SIEM={eval_meta.get('profile_select', {}).get('siem_type', 'none')}")
    lines.append(f"- **수행 도구**: {', '.join(eval_meta.get('selected_tools') or ['(없음)'])}")
    lines.append(f"- **제외 도구**: {', '.join(eval_meta.get('excluded_tools') or ['(없음)'])}")
    tgt = eval_meta.get("scan_targets") or {}
    if tgt:
        lines.append(f"- **스캔 대상**: " + " · ".join(f"{k}={v}" for k, v in tgt.items()))
    consent = eval_meta.get("scan_consent") or {}
    if consent:
        lines.append(f"- **승인자**: {consent.get('approver', '(미입력)')}")
        lines.append(f"- **시간대**: {consent.get('scheduled_window', '(미입력)')}")
        lines.append(f"- **강도**: {consent.get('intensity', '(미입력)')}")
        lines.append(f"- **비상 연락처**: {consent.get('emergency_contact', '(미입력)')}")
    lines.append("")

    eval_ver  = eval_meta.get("evaluation_version") or {}
    scope_as  = eval_meta.get("evaluation_scope_assets") or []
    data_cls  = eval_meta.get("data_classifications") or []
    reviewers = eval_meta.get("reviewers") or {}

    if eval_ver:
        lines.append("### 2.1 평가 대상 버전")
        lines.append("")
        if eval_ver.get("version_label"):
            lines.append(f"- **버전 라벨**: {eval_ver['version_label']}")
        if eval_ver.get("git_commit"):
            lines.append(f"- **Git commit**: `{eval_ver['git_commit']}`")
        if eval_ver.get("frontend_deployment"):
            lines.append(f"- **Frontend deployment**: {eval_ver['frontend_deployment']}")
        if eval_ver.get("backend_deployment"):
            lines.append(f"- **Backend deployment**: {eval_ver['backend_deployment']}")
        lines.append("")

    if scope_as:
        lines.append(f"### 2.2 평가 범위 자산 목록 ({len(scope_as)}건)")
        lines.append("")
        lines.append("| 포함 | 자산 | 값 |")
        lines.append("|---|---|---|")
        for a in scope_as:
            tag = "✓" if a.get("included") else "—"
            name = (a.get("name") or "").replace("|", "\\|")
            value = (a.get("value") or "").replace("|", "\\|")
            lines.append(f"| {tag} | {name} | `{value}` |")
        lines.append("")

    if data_cls:
        lines.append(f"### 2.3 데이터 등급 분류 ({len(data_cls)}건)")
        lines.append("")
        lines.append("| 민감도 | 데이터 항목 | 보관 위치 |")
        lines.append("|---|---|---|")
        for d in data_cls:
            name = (d.get("name") or "").replace("|", "\\|")
            loc = (d.get("storage_location") or "").replace("|", "\\|")
            lines.append(f"| {d.get('sensitivity', '')} | {name} | {loc} |")
        lines.append("")

    lines.append("## 3. 종합 결과")
    lines.append("")
    lines.append(f"- **총 항목**: {total}")
    lines.append(f"- **충족**: {pass_cnt}건 / **부분충족**: {partial_cnt}건 / "
                 f"**미충족**: {fail_cnt}건 / **평가불가**: {na_cnt}건")
    lines.append(f"- **종합 점수**: {round(session.total_score or 0.0, 2)} / 4.0  "
                 f"(레벨: {session.level or '-'})")
    lines.append("")

    lines.append(f"## 4. 논쟁 가능 항목 ({len(debate_results)}건)")
    lines.append("")
    lines.append("> 자동 수집된 metric / Evidence 관찰 내용 / 평가불가 사유를 정리합니다.  ")
    lines.append("> *리뷰어 의견* 칸은 비어 있으니, 검토 후 직접 작성해주세요.")
    lines.append("")

    if not debate_results:
        lines.append("(논쟁 가능 항목 없음 — 모든 항목이 충족 또는 미충족으로 명확히 판정)")
        lines.append("")
    else:
        current_pillar = None
        for dr, cl in debate_results:
            pillar = cl.pillar or "(Pillar 미상)"
            if pillar != current_pillar:
                lines.append(f"### {pillar}")
                lines.append("")
                current_pillar = pillar

            coll = coll_by_check.get(cl.check_id)
            evs = ev_by_check.get(cl.check_id, [])

            basis_lines: list[str] = []
            if coll:
                rj = coll.raw_json if isinstance(coll.raw_json, dict) else {}
                tool = coll.tool or "-"
                basis_lines.append(f"- **출처**: {'수동' if tool == '수동' else '자동'} ({tool})")
                if coll.metric_key:
                    basis_lines.append(
                        f"- **지표**: `{coll.metric_key}` = {coll.metric_value} (threshold {coll.threshold})"
                    )
                if coll.error:
                    basis_lines.append(f"- **오류**: {coll.error}")
                if rj.get("reason_code"):
                    basis_lines.append(
                        f"- **평가불가 사유**: `{rj.get('reason_code')}` — {rj.get('reason_label', '')}"
                    )
                if rj.get("issues"):
                    basis_lines.append(f"- **발견 이슈** ({len(rj['issues'])}건):")
                    for issue in rj["issues"][:5]:
                        basis_lines.append(f"  - {issue}")
                if rj.get("security_headers"):
                    sh = rj["security_headers"]
                    if isinstance(sh, dict) and sh.get("score") is not None:
                        basis_lines.append(
                            f"- **보안 헤더 종합**: {sh['score']:.2f}/1.0"
                        )

            for ev in evs:
                if ev.observed:
                    basis_lines.append(f"- **관찰 내용**: {ev.observed[:300]}")
                if ev.reason:
                    basis_lines.append(f"- **원인/근거**: {ev.reason[:300]}")
                if ev.original_filename:
                    basis_lines.append(f"- **증적 파일**: {ev.original_filename} ({ev.file_size or 0}B)")

            if not basis_lines:
                basis_lines.append("- (판정 근거 데이터 없음)")

            rec = (dr.recommendation or "").strip()

            lines.append(f"#### {cl.item_id} — {cl.item_name or '(항목명 미상)'} _({cl.maturity})_")
            lines.append("")
            lines.append(f"**결과**: `{dr.result}` (점수 {round(dr.score or 0.0, 2)})")
            lines.append("")
            lines.append("**판정 근거**:")
            lines.append("")
            for bl in basis_lines:
                lines.append(bl)
            lines.append("")
            if rec:
                lines.append(f"**권고**: {rec}")
                lines.append("")
            lines.append("**리뷰어 의견**:")
            lines.append("")
            lines.append("> _(여기에 검토 의견을 작성해주세요)_")
            lines.append("")
            lines.append("---")
            lines.append("")

    excluded = eval_meta.get("excluded_tools") or []
    if excluded:
        lines.append("## 5. 미사용 도구 (수행하지 않은 자동 진단)")
        lines.append("")
        for t in excluded:
            lines.append(f"- **{t}** — 사용자 환경 미선택 또는 비활성. 해당 도구 매핑 항목은 수동 진단으로 폴백됨.")
        lines.append("")

    lines.append("## 6. 리뷰어 서명")
    lines.append("")
    lines.append("> 수동 항목은 최소 2인 리뷰.")
    lines.append("")
    lines.append("| 역할 | 이름 | 검토 일자 | 서명 |")
    lines.append("|---|---|---|---|")
    lines.append(f"| App owner | {reviewers.get('app_owner', '')} |  |  |")
    lines.append(f"| Backend owner | {reviewers.get('backend_owner', '')} |  |  |")
    lines.append(f"| Cloud owner | {reviewers.get('cloud_owner', '')} |  |  |")
    lines.append(f"| Security reviewer | {reviewers.get('security_reviewer', '')} |  |  |")
    lines.append("")

    return "\n".join(lines)


@router.get("/decision-log/{session_id}")
async def download_decision_log(
    session_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """판정 로그 markdown 다운로드 — 가이드 §7 산출물 decision_log.md.

    부분충족·평가불가 항목의 판정 근거 + 리뷰어 의견(빈 칸) 정리.
    """
    session = db.query(DiagnosisSession).filter(
        DiagnosisSession.session_id == session_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
    assert_session_access(current_user, session)

    try:
        text = await asyncio.to_thread(_build_decision_log_md, session_id, db)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("[report] decision log build failed: %s", exc)
        raise HTTPException(status_code=500, detail="판정 로그 생성에 실패했습니다.")

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = f"decision-log-{session_id}-{today}.md"
    return StreamingResponse(
        io.BytesIO(text.encode("utf-8")),
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
