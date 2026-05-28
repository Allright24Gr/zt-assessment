from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import List, Optional
from datetime import datetime, timezone
from pathlib import Path
from copy import copy
import io
import os
import uuid
import logging
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

from database import get_db
from models import DiagnosisSession, Checklist, DiagnosisResult, Evidence, CollectedData, User
from routers.auth import get_current_user, assert_session_access

logger = logging.getLogger(__name__)
router = APIRouter()

# P1-7: 증적 파일 업로드 설정
EVIDENCE_STORAGE_DIR = os.getenv("EVIDENCE_STORAGE_DIR", "/var/lib/zt-assessment/evidence")
MAX_EVIDENCE_SIZE_MB = int(os.getenv("MAX_EVIDENCE_SIZE_MB", "10"))
_ALLOWED_EVIDENCE_MIMES = {
    "application/pdf":  "pdf",
    "image/png":        "png",
    "image/jpeg":       "jpg",
    "image/gif":        "gif",
    "image/webp":       "webp",
}

MATURITY_NUM = {"기존": 1, "초기": 2, "향상": 3, "최적화": 4}
VALID_RESULTS = {"충족", "부분충족", "미충족", "평가불가"}

# 담당자가 선택할 수 있는 4가지 직관 기호 — 자유 입력 차단.
# O = 충족, △ = 부분충족, X = 미충족, 평가불가 = 평가불가.
CHOICE_OPTIONS = ["O", "△", "X", "평가불가"]
CHOICE_TO_VERDICT = {
    "O":      "충족",
    "△":      "부분충족",
    "X":      "미충족",
    "평가불가": "평가불가",
}


def _normalize_result(v: str) -> str:
    """부분 충족 → 부분충족 등 공백 정규화."""
    return v.replace(" ", "") if v else "평가불가"


def resolve_verdict(user_choice: str) -> str:
    """담당자가 선택한 기호 → 판정 결과 변환.

    선택값이 정의되지 않은 값이면 '평가불가' 처리.
    """
    return CHOICE_TO_VERDICT.get((user_choice or "").strip(), "평가불가")


def _find_checklist(db: Session, item_no_str: str, maturity: str, question: str) -> Optional[Checklist]:
    """항목번호+성숙도+질문으로 Checklist 행을 찾는다.

    Excel 항목번호 예: "1.1.1 사용자 인벤토리" → prefix "1.1.1"
    DB item_id 형식: "{prefix}.{maturity_num}_{counter}" → "1.1.1.1_1"
    """
    if not item_no_str:
        return None
    prefix = item_no_str.strip().split()[0]
    mat_num = MATURITY_NUM.get(maturity, 0)
    if mat_num == 0:
        return None

    pattern = f"{prefix}.{mat_num}_%"
    candidates = db.query(Checklist).filter(Checklist.item_id.like(pattern)).all()
    if not candidates:
        return None

    q = (question or "").strip()
    if q:
        for c in candidates:
            if c.item_name and c.item_name.strip() == q:
                return c
    return candidates[0]


class ManualAnswer(BaseModel):
    check_id: str
    value: str
    evidence: str = ""


class ManualSubmitRequest(BaseModel):
    session_id: int
    answers: List[ManualAnswer]


@router.post("/upload")
async def manual_upload(
    session_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """수동 체크리스트 Excel 파일을 업로드해 DiagnosisResult를 일괄 생성한다.

    Excel 형식: manual-checklist.xlsx (manual_diagnosis + judgment_mapping 시트)
    """
    session = db.query(DiagnosisSession).filter(
        DiagnosisSession.session_id == session_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
    assert_session_access(current_user, session)

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="올바른 Excel(.xlsx) 파일이 아닙니다.")

    if "manual_diagnosis" not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail="manual_diagnosis 시트가 필요합니다.",
        )

    ws = wb["manual_diagnosis"]

    weight_map = {"충족": 1.0, "부분충족": 0.5, "미충족": 0.0, "평가불가": 0.0}
    saved = 0
    unmatched = 0
    skipped = 0

    for r in range(4, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 9)]
        m_id, category, item_no, maturity, question, _, choice, note = row

        # 카테고리 구분행·미입력 건너뜀
        if not m_id or str(m_id).startswith("▸") or not choice:
            skipped += 1
            continue

        m_id_str = str(m_id).strip()
        choice_str = str(choice).strip()
        note_str = str(note).strip() if note else ""

        # 신규 규칙: 담당자가 선택한 기호(O/△/X/평가불가) → 판정 직접 매핑.
        # 양식의 모든 행은 동일한 4선택 dropdown 으로 통일됨.
        verdict = resolve_verdict(choice_str)
        if verdict not in VALID_RESULTS:
            verdict = "평가불가"

        checklist = _find_checklist(db, str(item_no), str(maturity), str(question))
        if not checklist:
            unmatched += 1
            continue

        check_id = checklist.check_id
        score = checklist.maturity_score * weight_map.get(verdict, 0.0)

        # CollectedData upsert
        existing_cd = db.query(CollectedData).filter(
            CollectedData.session_id == session_id,
            CollectedData.check_id == check_id,
        ).first()
        if existing_cd:
            existing_cd.tool = "수동"
            existing_cd.metric_key = "manual_result"
            existing_cd.metric_value = weight_map.get(verdict, 0.0)
            existing_cd.threshold = 1.0
            existing_cd.raw_json = {"manual": True, "choice": choice_str, "note": note_str}
            existing_cd.error = None
            existing_cd.collected_at = datetime.now(timezone.utc)
        else:
            db.add(CollectedData(
                session_id=session_id,
                check_id=check_id,
                tool="수동",
                metric_key="manual_result",
                metric_value=weight_map.get(verdict, 0.0),
                threshold=1.0,
                raw_json={"manual": True, "choice": choice_str, "note": note_str},
                error=None,
            ))

        # DiagnosisResult upsert
        existing_dr = db.query(DiagnosisResult).filter(
            DiagnosisResult.session_id == session_id,
            DiagnosisResult.check_id == check_id,
        ).first()
        if existing_dr:
            existing_dr.result = verdict
            existing_dr.score = score
        else:
            db.add(DiagnosisResult(
                session_id=session_id,
                check_id=check_id,
                result=verdict,
                score=score,
                recommendation="",
            ))

        if note_str:
            db.add(Evidence(
                session_id=session_id,
                check_id=check_id,
                source="수동입력(Excel)",
                observed=note_str,
                location="",
                reason="",
                impact=None,
            ))

        saved += 1

    db.commit()
    return {
        "status":          "ok",
        "session_id":      session_id,
        "parsed_count":    saved,
        "unmatched_count": unmatched,
        "skipped_count":   skipped,
    }


@router.post("/submit")
def manual_submit(
    req: ManualSubmitRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    session = db.query(DiagnosisSession).filter(
        DiagnosisSession.session_id == req.session_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
    assert_session_access(current_user, session)

    weight_map = {"충족": 1.0, "부분충족": 0.5, "미충족": 0.0, "평가불가": 0.0}
    saved = 0

    for ans in req.answers:
        result = ans.value if ans.value in VALID_RESULTS else "평가불가"
        evidence_text = ans.evidence or None

        checklist = db.query(Checklist).filter(Checklist.item_id == ans.check_id).first()
        if not checklist or checklist.diagnosis_type != "수동":
            continue

        check_id = checklist.check_id

        existing_cd = db.query(CollectedData).filter(
            CollectedData.session_id == req.session_id,
            CollectedData.check_id == check_id,
        ).first()
        if existing_cd:
            existing_cd.tool = "수동"
            existing_cd.metric_key = "manual_result"
            existing_cd.metric_value = weight_map.get(result, 0.0)
            existing_cd.threshold = 1.0
            existing_cd.raw_json = {"manual": True, "evidence": evidence_text}
            existing_cd.error = None
            existing_cd.collected_at = datetime.now(timezone.utc)
        else:
            db.add(CollectedData(
                session_id=req.session_id,
                check_id=check_id,
                tool="수동",
                metric_key="manual_result",
                metric_value=weight_map.get(result, 0.0),
                threshold=1.0,
                raw_json={"manual": True, "evidence": evidence_text},
                error=None,
            ))

        existing = db.query(DiagnosisResult).filter(
            DiagnosisResult.session_id == req.session_id,
            DiagnosisResult.check_id == check_id,
        ).first()
        if existing:
            existing.result = result
            existing.score = checklist.maturity_score * weight_map.get(result, 0.0)
            existing.recommendation = ""
        else:
            db.add(DiagnosisResult(
                session_id=req.session_id,
                check_id=check_id,
                result=result,
                score=checklist.maturity_score * weight_map.get(result, 0.0),
                recommendation="",
            ))

        if evidence_text:
            db.add(Evidence(
                session_id=req.session_id,
                check_id=check_id,
                source="수동입력",
                observed=evidence_text,
                location="",
                reason="",
                impact=None,
            ))

        saved += 1

    db.commit()

    return {
        "status": "ok",
        "session_id": req.session_id,
        "submitted_count": saved,
    }


@router.get("/template")
def download_template(current_user: User = Depends(get_current_user)):
    """사용자에게 배포할 빈 manual-checklist.xlsx 템플릿을 반환한다.

    정적 양식 — xlsx 원래 수동 진단 98건만 포함. 자동 폴백 항목은 미포함이므로
    세션이 있다면 GET /template/{session_id} 동적 양식 사용을 권장.
    """
    p = _find_base_template()
    if p:
        return FileResponse(
            path=str(p),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="manual-checklist.xlsx",
        )
    raise HTTPException(status_code=404, detail="템플릿 파일을 찾을 수 없습니다.")


def _find_base_template() -> Optional[Path]:
    """배포 환경 (/app/...) 과 로컬 dev 환경 양쪽에서 base xlsx 위치 탐색."""
    for p in (Path("/app/manual-checklist.xlsx"),
              Path(__file__).parent.parent / "manual-checklist.xlsx"):
        if p.exists():
            return p
    return None


# ─── 세션별 동적 양식 생성 ──────────────────────────────────────────────────
# T-Markov 같은 SaaS 환경(Keycloak/Wazuh 미사용)에서는 자동 항목이 수동으로
# 폴백되는데, 정적 양식엔 그 폴백 항목이 없다. 세션의 profile_select 와
# selected_tools 를 기반으로 폴백 항목까지 포함한 양식을 동적 생성한다.

_IDP_AUTO_TOOLS_FOR_FALLBACK = {"keycloak", "supabase", "entra"}
_SIEM_AUTO_TOOLS_FOR_FALLBACK = {"wazuh"}


def _resolve_fallback_tools(session: DiagnosisSession) -> set[str]:
    """세션 환경에서 자동 폴백되는 도구 집합 산출.

    예: profile_select={idp_type: 'google_workspace', siem_type: 'none'}
        → {'keycloak', 'entra', 'wazuh'} 모두 폴백(자동 항목이 수동으로 떨어짐).
    """
    extra = session.extra if isinstance(session.extra, dict) else {}
    ps = extra.get("profile_select") if isinstance(extra.get("profile_select"), dict) else {}
    idp_sel = (ps.get("idp_type") or "").lower()
    siem_sel = (ps.get("siem_type") or "").lower()

    fallback: set[str] = set()
    # IdP 자동 도구 중 사용자가 선택하지 않은 것 → 폴백
    if idp_sel:
        for t in _IDP_AUTO_TOOLS_FOR_FALLBACK:
            if idp_sel != t:
                fallback.add(t)
    if siem_sel:
        for t in _SIEM_AUTO_TOOLS_FOR_FALLBACK:
            if siem_sel != t:
                fallback.add(t)
    return fallback


# 폴백 항목도 본 양식과 동일하게 O/△/X/평가불가 4선택 dropdown 사용.
# (choice → verdict 매핑은 CHOICE_TO_VERDICT 참조)
_FALLBACK_CHOICES = [(c, CHOICE_TO_VERDICT[c]) for c in CHOICE_OPTIONS]


_MATURITY_NAME_FROM_NUM = {1: "기존", 2: "초기", 3: "향상", 4: "최적화"}


def _category_prefix(item_id: str) -> str:
    """item_id 'X.Y.Z.N_M' → 카테고리 prefix 'X.Y.Z'."""
    if not item_id:
        return ""
    base = item_id.split("_", 1)[0]
    parts = base.split(".")
    return ".".join(parts[:3]) if len(parts) >= 3 else ""


def _maturity_num_from_item_id(item_id: str) -> int:
    """item_id 'X.Y.Z.N_M' → maturity 숫자 N (1=기존, 2=초기, 3=향상, 4=최적화)."""
    if not item_id:
        return 0
    base = item_id.split("_", 1)[0]
    parts = base.split(".")
    if len(parts) < 4:
        return 0
    try:
        return int(parts[3])
    except (ValueError, TypeError):
        return 0


def _analyze_auto_max_level(session_id: int, db: Session) -> dict[str, int]:
    """세션의 자동 진단 결과를 분석해 카테고리(prefix) 별 충족된 최고 단계 산출.

    KISA ZT 가이드라인의 보수적 평가 원칙: 상위 단계가 충족되면 하위 단계는 자동
    충족된 것으로 본다. 따라서 수동 양식에서는 자동으로 충족된 최고 단계 *이하* 행을
    모두 제외해야 한다.

    반환: {"1.1.1": 3, "1.2.1": 0, ...}
      · 키: 카테고리 prefix (3-segment item_id)
      · 값: 자동 진단에서 "충족" 으로 판정된 가장 높은 maturity 숫자 (1~4).
            충족된 게 없으면 키 자체가 없음 → caller 는 .get(prefix, 0) 사용.
    """
    rows = (
        db.query(DiagnosisResult, Checklist)
        .join(Checklist, DiagnosisResult.check_id == Checklist.check_id)
        .filter(DiagnosisResult.session_id == session_id)
        .all()
    )
    auto_max: dict[str, int] = {}
    for dr, cl in rows:
        if cl.diagnosis_type != "자동":
            continue
        if dr.result != "충족":
            continue
        prefix = _category_prefix(cl.item_id or "")
        mat_num = _maturity_num_from_item_id(cl.item_id or "")
        if not prefix or not mat_num:
            continue
        prev = auto_max.get(prefix, 0)
        if mat_num > prev:
            auto_max[prefix] = mat_num
    return auto_max


def _auto_result_distribution(session_id: int, db: Session) -> dict[str, dict]:
    """카테고리 prefix 별 자동 결과 분포 + 한국어 카테고리 이름 산출 (요약 시트용).

    반환: {
        "1.1.1": {
            "category_label": "1.1.1 사용자 인벤토리",
            "max_satisfied":  3,   # 자동 최고 충족 단계 (없으면 0)
            "counts": {"충족": 2, "부분충족": 0, "미충족": 1, "평가불가": 0},
        },
        ...
    }
    """
    rows = (
        db.query(DiagnosisResult, Checklist)
        .join(Checklist, DiagnosisResult.check_id == Checklist.check_id)
        .filter(DiagnosisResult.session_id == session_id)
        .all()
    )
    summary: dict[str, dict] = {}
    for dr, cl in rows:
        if cl.diagnosis_type != "자동":
            continue
        prefix = _category_prefix(cl.item_id or "")
        if not prefix:
            continue
        rec = summary.setdefault(prefix, {
            "category_label": cl.category or prefix,
            "max_satisfied":  0,
            "counts":         {"충족": 0, "부분충족": 0, "미충족": 0, "평가불가": 0},
        })
        # 가장 의미 있는 라벨 (Checklist.category) 채택
        if cl.category and len(cl.category) > len(rec["category_label"]):
            rec["category_label"] = cl.category
        result_key = _normalize_result(dr.result or "")
        if result_key in rec["counts"]:
            rec["counts"][result_key] += 1
        if result_key == "충족":
            mat_num = _maturity_num_from_item_id(cl.item_id or "")
            if mat_num > rec["max_satisfied"]:
                rec["max_satisfied"] = mat_num
    return summary


# 카테고리(Pillar) 정렬 순서 — 기존 양식과 동일하게 6 Pillar 순으로.
_PILLAR_ORDER = [
    "식별자 및 신원",
    "기기 및 엔드포인트",
    "네트워크",
    "시스템",
    "애플리케이션 및 워크로드",
    "데이터",
]


def _pillar_sort_key(pillar: str) -> int:
    try:
        return _PILLAR_ORDER.index(pillar)
    except ValueError:
        return len(_PILLAR_ORDER)


def _build_session_template_xlsx(session: DiagnosisSession, db: Session) -> bytes:
    """세션 기반 동적 xlsx 생성 — 기존 정적 양식에 폴백 항목 추가 +
    공개 URL 자동 점검 결과를 비고/부록 시트에 자동 채움.

    동작:
      1) base manual-checklist.xlsx 로드.
      2) 폴백 도구의 Checklist 항목을 DB 에서 조회 (이미 수동 양식에 있는 항목은 제외).
      3) 공개 URL 자동 점검 (web_evidence_collector) — HTTP 헤더/DNS/TLS/공개노출/GitHub repo
      4) manual_diagnosis 시트 끝에 카테고리 구분행 + 데이터 행 추가.
         각 행의 비고/증적메모 컬럼에 Pillar 관련 자동 점검 요약 미리 채움.
      5) judgment_mapping 시트 끝에 동일 M_id 의 4선택지(_FALLBACK_CHOICES) 추가.
      6) 새 시트 "외부 자동 점검 결과" — 자동 점검 raw 결과 정리 (참고용).
      7) bytes 로 직렬화.
    """
    base = _find_base_template()
    if not base:
        raise HTTPException(status_code=500, detail="base 양식 파일을 찾을 수 없습니다.")

    wb = openpyxl.load_workbook(base)
    ws_diag = wb["manual_diagnosis"]
    ws_judg = wb["judgment_mapping"]

    # ── 자동 진단 결과 분석 (보수적 단계 평가) ────────────────────────────────
    # 카테고리별 자동 충족된 최고 단계 ≥ 행 단계인 행은 양식에서 제외한다.
    # ex) 1.1.1 카테고리에서 "향상(3)" 까지 자동 충족 → 1.1.1 의 기존/초기/향상 행 모두 제외.
    auto_max = _analyze_auto_max_level(session.session_id, db)
    auto_dist = _auto_result_distribution(session.session_id, db)

    # ── 양식 R2 안내문에 보수적 평가 안내 추가 ────────────────────────────────
    notice_cell = ws_diag.cell(2, 1)
    base_notice = str(notice_cell.value or "")
    extra_notice = " ※ 자동 진단에서 충족된 항목은 보수적 평가 원칙에 따라 양식에서 자동 제외되었습니다 (해당 카테고리 자동 결과 요약 시트 참고)."
    if extra_notice not in base_notice:
        notice_cell.value = base_notice + extra_notice

    # ── 공개 URL 자동 점검 (실패해도 양식 생성은 진행) ───────────────────────
    extra = session.extra if isinstance(session.extra, dict) else {}
    scan_targets = extra.get("scan_targets") if isinstance(extra.get("scan_targets"), dict) else {}
    nmap_target = (scan_targets.get("nmap") or "").strip()
    trivy_target = (scan_targets.get("trivy") or "").strip()
    # Trivy target 이 GitHub 형식이면 repo 분석에 활용
    github_ref = trivy_target if trivy_target and ("github.com" in trivy_target or "/" in trivy_target and ":" not in trivy_target) else ""

    evidence: dict = {}
    if nmap_target or github_ref:
        try:
            from collectors import web_evidence_collector as _wec
            evidence = _wec.collect_public_evidence(
                nmap_target=nmap_target,
                github_repo=github_ref,
                timeout=8.0,
            )
        except Exception as exc:
            logger.warning("[manual] web evidence collection failed: %s", exc)
            evidence = {"error": str(exc)}

    # ── 자동 충족된 행 제거 (보수적 단계 평가) ────────────────────────────────
    # 역순 순회로 행 번호 안 꼬이게 처리. 데이터 행 (M### 시작) 만 검사 대상.
    removed_rows_info: list[tuple[str, str]] = []  # (prefix, maturity) — 로그/디버그용
    for r in range(ws_diag.max_row, 3, -1):
        mid = ws_diag.cell(r, 1).value
        if not mid or not isinstance(mid, str) or not mid.startswith("M"):
            continue
        item_no_cell = ws_diag.cell(r, 3).value
        maturity_cell = ws_diag.cell(r, 4).value
        if not item_no_cell or not maturity_cell:
            continue
        prefix = str(item_no_cell).split()[0]
        maturity = str(maturity_cell).strip()
        mat_num = MATURITY_NUM.get(maturity, 0)
        if not prefix or not mat_num:
            continue
        category_max = auto_max.get(prefix, 0)
        if category_max >= mat_num:
            ws_diag.delete_rows(r, 1)
            removed_rows_info.append((prefix, maturity))

    # 빈 카테고리 구분행 (▸ ...) 정리 — 직후가 또 ▸ 이거나 시트 끝이면 제거.
    for r in range(ws_diag.max_row, 3, -1):
        v = ws_diag.cell(r, 1).value
        if not v or not isinstance(v, str) or not v.startswith("▸"):
            continue
        nxt = ws_diag.cell(r + 1, 1).value if r + 1 <= ws_diag.max_row else None
        if not nxt or (isinstance(nxt, str) and nxt.startswith("▸")):
            ws_diag.delete_rows(r, 1)

    # M_id 재번호 — 빈자리 없이 M001~ 연속 번호.
    next_m_for_renumber = 1
    for r in range(4, ws_diag.max_row + 1):
        v = ws_diag.cell(r, 1).value
        if v and isinstance(v, str) and v.startswith("M"):
            ws_diag.cell(r, 1).value = f"M{next_m_for_renumber:03d}"
            next_m_for_renumber += 1

    # 기존 양식에 이미 포함된 (item_no_prefix, maturity) 조합 — 폴백 중복 추가 방지.
    existing_keys: set[tuple[str, str]] = set()
    for r in range(4, ws_diag.max_row + 1):
        mid = ws_diag.cell(r, 1).value
        if not mid or str(mid).startswith("▸"):
            continue
        item_no = str(ws_diag.cell(r, 3).value or "").split()[0] if ws_diag.cell(r, 3).value else ""
        maturity = str(ws_diag.cell(r, 4).value or "").strip()
        if item_no and maturity:
            existing_keys.add((item_no, maturity))

    # 양식이 완전히 비었으면 안내 한 줄 (행 4 = 첫 카테고리 위치).
    has_any_data = next_m_for_renumber > 1
    if not has_any_data:
        info_row = 4
        ic = ws_diag.cell(info_row, 1)
        ic.value = "✓ 모든 자동 진단이 충족되어 수동 보완 항목이 없습니다. (자동 결과 요약 시트 참고)"
        ic.font = Font(name="Arial", size=11, bold=True, color="FF0F766E")
        ic.fill = PatternFill("solid", fgColor="FFE7F5F2")
        ic.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_diag.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=8)

    fallback_tools = _resolve_fallback_tools(session)
    # instructions 시트에도 보수적 평가 안내 추가 (멱등).
    _augment_instructions_sheet(wb)

    # 자동 진단에서 평가불가로 떨어진 항목들도 수동 입력으로 보완 — 결과 페이지에
    # "평가 안 됨" 공백을 남기지 않기 위함. KISA 보수적 평가 원칙은 동일하게 적용
    # (해당 카테고리의 상위 단계가 자동 충족이면 하위 단계 평가불가 항목은 제외).
    unavailable_check_ids: set[int] = set()
    unavailable_rows = (
        db.query(DiagnosisResult, Checklist)
        .join(Checklist, DiagnosisResult.check_id == Checklist.check_id)
        .filter(
            DiagnosisResult.session_id == session.session_id,
            DiagnosisResult.result == "평가불가",
        )
        .all()
    )
    unavailable_checklist: list[Checklist] = []
    for _dr, cl in unavailable_rows:
        if cl.check_id in unavailable_check_ids:
            continue
        unavailable_check_ids.add(cl.check_id)
        unavailable_checklist.append(cl)

    if not fallback_tools and not unavailable_checklist:
        # 폴백 없음 + 평가불가 없음 — 자동 결과 요약/외부 점검 시트만 부착 후 종료.
        _append_auto_summary_sheet(wb, auto_dist)
        if evidence and not evidence.get("error"):
            _append_evidence_sheet(wb, evidence)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # 폴백 도구 매핑된 Checklist 항목 조회 + 자동 평가불가 항목 합치기 (중복 제거).
    rows: list[Checklist] = []
    seen_check_ids: set[int] = set()
    if fallback_tools:
        for cl in db.query(Checklist).filter(
            Checklist.tool.in_(sorted(fallback_tools))
        ).all():
            if cl.check_id not in seen_check_ids:
                seen_check_ids.add(cl.check_id)
                rows.append(cl)
    for cl in unavailable_checklist:
        if cl.check_id not in seen_check_ids:
            seen_check_ids.add(cl.check_id)
            rows.append(cl)
    # 카테고리·item_id 순으로 정렬
    rows.sort(key=lambda c: (_pillar_sort_key(c.pillar or ""), c.item_id or ""))

    # 기존 양식에 이미 있는 항목 제외 + 자동 충족된 단계 제외 (보수적 평가).
    new_items = []
    for cl in rows:
        prefix = (cl.item_id or "").split("_", 1)[0]  # "1.1.1.1"
        item_no_prefix = ".".join(prefix.split(".")[:3])  # "1.1.1"
        key = (item_no_prefix, cl.maturity or "")
        if key in existing_keys:
            continue
        # 자동에서 같은 카테고리의 상위(또는 동등) 단계가 충족되어 있으면 폴백도 생략.
        mat_num = MATURITY_NUM.get(cl.maturity or "", 0)
        if mat_num and auto_max.get(item_no_prefix, 0) >= mat_num:
            removed_rows_info.append((item_no_prefix, cl.maturity or ""))
            continue
        new_items.append((cl, item_no_prefix))

    if not new_items:
        # 자동 결과 요약 시트는 폴백 없이도 부착해서 담당자가 확인 가능.
        _append_auto_summary_sheet(wb, auto_dist)
        if evidence and not evidence.get("error"):
            _append_evidence_sheet(wb, evidence)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # 다음 M_id 번호 산출
    max_m_num = 0
    for r in range(4, ws_diag.max_row + 1):
        v = ws_diag.cell(r, 1).value
        if v and isinstance(v, str) and v.startswith("M"):
            try:
                max_m_num = max(max_m_num, int(v[1:]))
            except ValueError:
                pass
    next_m = max_m_num + 1

    # 행 스타일 캐시 — base 양식의 데이터 행 1개를 샘플로 사용 (Row 5)
    sample_row = 5
    cell_styles = {}
    for col in range(1, 9):
        sc = ws_diag.cell(sample_row, col)
        cell_styles[col] = {
            "font":      copy(sc.font),
            "fill":      copy(sc.fill),
            "border":    copy(sc.border),
            "alignment": copy(sc.alignment),
        }

    # 카테고리 구분행 스타일 (Row 4)
    cat_styles = {}
    for col in range(1, 9):
        sc = ws_diag.cell(4, col)
        cat_styles[col] = {
            "font":      copy(sc.font),
            "fill":      copy(sc.fill),
            "border":    copy(sc.border),
            "alignment": copy(sc.alignment),
        }

    # 폴백/평가불가 항목을 Pillar 별로 그룹화 + Pillar 내부는 item_id 순서.
    by_pillar: dict[str, list] = {}
    for cl, item_no_prefix in new_items:
        by_pillar.setdefault(cl.pillar or "기타", []).append((cl, item_no_prefix))
    for pillar in by_pillar:
        by_pillar[pillar].sort(key=lambda t: t[0].item_id or "")

    # 사전 M_id 할당 — xlsx 최종 visual order(식별자→기기→...→데이터) 기준 연속 번호.
    # 이렇게 해야 reverse 삽입을 해도 보기에는 위→아래 M### 가 단조 증가.
    pre_assigned_mids: list[str] = []
    pillar_order_keys = [p for p in _PILLAR_ORDER if by_pillar.get(p)]
    extra_pillars = sorted([p for p in by_pillar if p not in _PILLAR_ORDER])
    for pillar in pillar_order_keys + extra_pillars:
        for _ in by_pillar[pillar]:
            pre_assigned_mids.append(f"M{next_m:03d}")
            next_m += 1

    # 기존 sheet 에서 Pillar 별 데이터 행 마지막 위치 스캔.
    # ▸ 로 시작하는 카테고리 구분행을 기준으로 섹션 경계 파악.
    pillar_section_end: dict[str, int] = {}  # pillar -> 그 pillar 의 마지막 데이터 행
    current_pillar: Optional[str] = None
    section_start_row = 4
    for r in range(4, ws_diag.max_row + 1):
        v = ws_diag.cell(r, 1).value
        if v and isinstance(v, str) and v.startswith("▸"):
            if current_pillar:
                pillar_section_end[current_pillar] = section_start_row - 1
            current_pillar = v.replace("▸", "").strip()
            # 마지막 데이터 행 = 다음 ▸ 의 한 줄 위. 일단 r 로 두고 다음 ▸ 만나면 갱신.
            section_start_row = r + 1
            pillar_section_end[current_pillar] = ws_diag.max_row  # 일단 끝까지로 두고 다음 섹션 발견 시 갱신
        elif v and isinstance(v, str) and v.startswith("M") and current_pillar:
            pillar_section_end[current_pillar] = r

    # 폴백 강조용 스타일 (M_id 셀 + 판정 선택 셀 두 군데에만 옅은 빨강 fill).
    fallback_mid_font = Font(name="Arial", size=10, bold=True, color="FFB91C1C")
    fallback_mid_fill = PatternFill("solid", fgColor="FFFEE2E2")

    new_mapping_rows: list[tuple] = []  # (m_id, pillar, item_no_full, maturity, choice, verdict)
    inserted_row_ranges: list[tuple[int, int]] = []  # 드롭다운 적용용 (start, end inclusive)

    # 본문: reverse pillar order 로 삽입 → earlier pillar 위치 안 흔들림.
    # M_id 는 pre_assigned_mids 에서 visual order 로 미리 잘라둠.
    visual_order = pillar_order_keys + extra_pillars
    mid_cursor = 0
    pillar_mid_slice: dict[str, list[str]] = {}
    for pillar in visual_order:
        cnt = len(by_pillar[pillar])
        pillar_mid_slice[pillar] = pre_assigned_mids[mid_cursor:mid_cursor + cnt]
        mid_cursor += cnt

    for pillar in reversed(visual_order):
        items = by_pillar[pillar]
        mids = pillar_mid_slice[pillar]
        if not items:
            continue

        # Pillar 별 자동 점검 요약 (한 번만 계산해서 같은 Pillar 행에 공유)
        pillar_evidence_text = ""
        if evidence and not evidence.get("error"):
            try:
                from collectors import web_evidence_collector as _wec
                pillar_evidence_text = _wec.summarize_for_pillar(evidence, pillar)
            except Exception:
                pillar_evidence_text = ""

        # 삽입 위치: 기존 pillar 섹션의 마지막 데이터 행 바로 다음.
        # 기존에 그 pillar 가 없으면 시트 맨 끝에 새 카테고리 구분행과 함께 추가.
        if pillar in pillar_section_end:
            insert_at = pillar_section_end[pillar] + 1
            # N개 행 삽입 (기존 행들 아래로 밀림)
            ws_diag.insert_rows(insert_at, amount=len(items))
            data_start = insert_at
        else:
            # 새 pillar — 끝에 카테고리 구분행 + 데이터 행 모두 추가
            data_start = ws_diag.max_row + 2
            cat_cell = ws_diag.cell(data_start, 1)
            cat_cell.value = f"▸  {pillar}"
            ws_diag.merge_cells(start_row=data_start, start_column=1, end_row=data_start, end_column=8)
            for col in range(1, 9):
                c = ws_diag.cell(data_start, col)
                cs = cat_styles.get(col, {})
                if cs.get("font"):      c.font      = copy(cs["font"])
                if cs.get("fill"):      c.fill      = copy(cs["fill"])
                if cs.get("border"):    c.border    = copy(cs["border"])
                if cs.get("alignment"): c.alignment = copy(cs["alignment"])
            data_start += 1

        # 데이터 행 채우기
        first_new = data_start
        for i, (cl, item_no_prefix) in enumerate(items):
            r = data_start + i
            m_id = mids[i]
            item_no_full = (cl.category or item_no_prefix)
            row_vals = [
                m_id,
                pillar,
                item_no_full,
                cl.maturity or "",
                cl.item_name or "",
                cl.criteria or "",
                None,                          # ★ 담당자 선택 (필수) — 빈 칸
                pillar_evidence_text or None,  # 비고/증적메모
            ]
            for col, val in enumerate(row_vals, start=1):
                c = ws_diag.cell(r, col)
                c.value = val
                cs = cell_styles.get(col, {})
                if cs.get("font"):      c.font      = copy(cs["font"])
                if cs.get("fill"):      c.fill      = copy(cs["fill"])
                if cs.get("border"):    c.border    = copy(cs["border"])
                if cs.get("alignment"): c.alignment = copy(cs["alignment"])
            # M_id 셀에 폴백 강조 (옅은 빨강 배경 + 빨강 볼드)
            mid_cell = ws_diag.cell(r, 1)
            mid_cell.font = copy(fallback_mid_font)
            mid_cell.fill = copy(fallback_mid_fill)

            for choice in CHOICE_OPTIONS:
                verdict = CHOICE_TO_VERDICT[choice]
                new_mapping_rows.append(
                    (m_id, pillar, item_no_full, cl.maturity or "", choice, verdict)
                )
        inserted_row_ranges.append((first_new, first_new + len(items) - 1))

    # judgment_mapping 시트에 폴백 매핑 추가
    judg_append = ws_judg.max_row + 1
    for tup in new_mapping_rows:
        for col, val in enumerate(tup, start=1):
            ws_judg.cell(judg_append, col).value = val
        judg_append += 1

    # 전체 M_id 재번호 — 삽입 후 위→아래 순서로 M001 ~ Mxxx 단조 증가하도록.
    renum = 1
    for r in range(4, ws_diag.max_row + 1):
        v = ws_diag.cell(r, 1).value
        if v and isinstance(v, str) and v.startswith("M"):
            ws_diag.cell(r, 1).value = f"M{renum:03d}"
            renum += 1

    # 드롭다운(데이터 검증) 추가 — ★ 담당자 선택 컬럼(G) 모든 새 행에 일괄 적용.
    choice_str = ",".join(CHOICE_OPTIONS)
    dv = DataValidation(
        type="list",
        formula1=f'"{choice_str}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="선택값 확인",
        error="O / △ / X / 평가불가 중에서 선택해주세요.",
    )
    # 모든 신규 삽입 구간을 DV 에 등록
    for (s, e) in inserted_row_ranges:
        dv.add(f"G{s}:G{e}")
    ws_diag.add_data_validation(dv)
    # 아래 기존 코드는 사용 안 함 — 변수만 정의해 호환성 유지.
    first_data_row = inserted_row_ranges[0][0] if inserted_row_ranges else (ws_diag.max_row + 1)
    # notice_row 는 더 이상 안 그림 — 폴백 강조는 M_id 셀 빨강 배경으로 대체.
    notice_row = first_data_row - 1
    # 위에서 이미 dv 추가했으므로 여기서는 호환성 유지를 위한 노옵 (기존 변수 사용).
    _ = first_data_row  # noqa: F841

    # ── 가독성: 열 너비 + 행 높이 자동 조정 ───────────────────────────────
    # 기존 base 양식의 컬럼 너비가 일부 짧아 글자가 잘려보이는 문제 해소.
    # A=M_id / B=Pillar / C=항목번호·이름 / D=성숙도 / E=세부질문 / F=판정기준 / G=선택 / H=비고
    column_widths = {
        "A": 8,    # M001 ~ M999
        "B": 18,   # Pillar 이름 (식별자 및 신원 등)
        "C": 28,   # 항목번호·이름
        "D": 10,   # 성숙도 (기존/초기/향상/최적화)
        "E": 60,   # 세부 질문 — 가장 길게
        "F": 50,   # 판정 기준
        "G": 14,   # 선택 (O/△/X/평가불가)
        "H": 40,   # 비고/증적메모
    }
    for col_letter, width in column_widths.items():
        ws_diag.column_dimensions[col_letter].width = width

    # 데이터 행은 wrap_text 가 켜져 있으므로 높이 자동 — 헤더와 일부 행만 명시.
    # 모든 데이터 행에 충분한 높이 보장 (긴 세부질문/판정기준 2~3줄 wrap 가능).
    for r in range(4, ws_diag.max_row + 1):
        v = ws_diag.cell(r, 1).value
        if v and isinstance(v, str) and v.startswith("M"):
            # 데이터 행: 32pt 정도면 wrap 2~3줄 무리 없음
            ws_diag.row_dimensions[r].height = 32

    # ── 부록 시트: 자동 결과 요약 + 외부 자동 점검 결과 (참고용) ───────────
    _append_auto_summary_sheet(wb, auto_dist)
    if evidence and not evidence.get("error"):
        _append_evidence_sheet(wb, evidence)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _augment_instructions_sheet(wb: openpyxl.Workbook) -> None:
    """instructions 시트에 보수적 평가 안내 한 줄 추가 (멱등).

    이미 추가되어 있으면 아무 일도 하지 않는다. 시트 자체가 없으면 무시.
    """
    if "instructions" not in wb.sheetnames:
        return
    ws_i = wb["instructions"]
    marker = "자동 진단에서 충족된 항목은 보수적 평가 원칙"
    for r in range(1, ws_i.max_row + 1):
        for c in range(1, 5):
            v = ws_i.cell(r, c).value
            if isinstance(v, str) and marker in v:
                return  # 이미 안내됨
    append_at = ws_i.max_row + 2
    c1 = ws_i.cell(append_at, 1, "자동 처리")
    c1.font = Font(name="Arial", size=10, bold=True, color="FF0F766E")
    c2 = ws_i.cell(append_at, 2,
        "자동 진단에서 충족된 항목은 보수적 평가 원칙(상위 충족 → 하위 자동 충족)에 따라 "
        "양식에서 자동 제외되었습니다. 어떤 자동 결과로 어떤 단계가 빠졌는지는 "
        "'자동 결과 요약' 시트에서 확인할 수 있습니다."
    )
    c2.alignment = Alignment(wrap_text=True, vertical="top")


def _append_auto_summary_sheet(wb: openpyxl.Workbook, auto_dist: dict[str, dict]) -> None:
    """양식에 '자동 결과 요약' 시트 추가 — 카테고리별 자동 충족 단계 + 결과 분포.

    담당자가 어떤 자동 결과 때문에 어떤 단계가 양식에서 빠졌는지 한눈에 보도록.
    auto_dist 가 비어있어도 헤더만 있는 빈 시트는 만들지 않는다.
    """
    if not auto_dist:
        return
    # 중복 호출 방지 — 같은 이름 시트가 이미 있으면 건너뜀.
    sheet_name = "자동 결과 요약"
    if sheet_name in wb.sheetnames:
        return
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40

    bold = Font(name="Arial", size=11, bold=True, color="FF1E3A5F")
    header_fill = PatternFill("solid", fgColor="FFE7F5F2")
    title_font = Font(name="Arial", size=12, bold=True, color="FF1E3A5F")
    title_fill = PatternFill("solid", fgColor="FFFEF3C7")

    tc = ws.cell(1, 1, "자동 진단 결과 요약 — 보수적 평가에 의해 양식에서 제외된 범위 확인용")
    tc.font = title_font
    tc.fill = title_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    tc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["카테고리(prefix)", "카테고리 이름", "자동 최고 충족 단계", "자동 결과 분포"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(3, col, h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # 정렬: prefix 사전순 (1.1.1 → 1.1.2 → ...)
    row = 4
    def _prefix_key(p: str) -> tuple:
        try:
            return tuple(int(x) for x in p.split("."))
        except Exception:
            return (999,)
    for prefix in sorted(auto_dist.keys(), key=_prefix_key):
        rec = auto_dist[prefix]
        max_n = rec.get("max_satisfied", 0)
        max_label = _MATURITY_NAME_FROM_NUM.get(max_n, "(없음)") if max_n else "(없음)"
        counts = rec.get("counts", {})
        dist_text = " / ".join(
            f"{k} {counts.get(k, 0)}"
            for k in ("충족", "부분충족", "미충족", "평가불가")
        )
        ws.cell(row, 1, prefix)
        ws.cell(row, 2, rec.get("category_label") or prefix)
        ws.cell(row, 3, max_label)
        ws.cell(row, 4, dist_text)
        for col in range(1, 5):
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
        row += 1


def _append_evidence_sheet(wb: openpyxl.Workbook, evidence: dict) -> None:
    """양식에 '외부 자동 점검 결과' 시트 추가 — 양식 비고 자동 채움의 raw 결과 노출.

    사용자가 양식 작성 시 비고에 적힌 한 줄 요약 외에 자세한 발견 사항을 참고할 수 있도록.
    """
    ws = wb.create_sheet("외부 자동 점검 결과")
    bold = Font(name="Arial", size=11, bold=True, color="FF1E3A5F")
    header_fill = PatternFill("solid", fgColor="FFE7F5F2")
    section = Font(name="Arial", size=10, bold=True, color="FF0F766E")

    def header(text: str):
        nonlocal row
        c = ws.cell(row, 1, text)
        c.font = bold
        c.fill = header_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

    def section_title(text: str):
        nonlocal row
        c = ws.cell(row, 1, text)
        c.font = section
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

    def kv(k: str, v):
        nonlocal row
        ws.cell(row, 1, k).font = Font(name="Arial", size=9, color="FF617087")
        ws.cell(row, 2, str(v) if v is not None else "")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 40

    row = 1
    header("Readyz-T 자동 점검 결과 (참고)")
    kv("생성 시각", evidence.get("generated_at", ""))
    kv("점검 대상 (Nmap target)", evidence.get("nmap_target") or "(미지정)")
    kv("점검 대상 (GitHub repo)", evidence.get("github_repo") or "(미지정)")
    row += 1

    # HTTP 헤더
    http = evidence.get("http_headers") or {}
    if http and not http.get("error"):
        section_title("◆ HTTP 보안 헤더")
        kv("URL", http.get("url"))
        kv("Status", http.get("status"))
        kv("종합 점수", f"{http.get('score', 0):.2f} / 1.0")
        for k, v in (http.get("assessment") or {}).items():
            kv(f"  · {k}", v)
        for i, issue in enumerate(http.get("issues") or [], 1):
            kv(f"이슈 {i}", issue)
        row += 1

    # DNS
    dns = evidence.get("dns") or {}
    if dns and not dns.get("error"):
        section_title("◆ DNS 보안 레코드")
        kv("도메인", dns.get("domain"))
        kv("종합 점수", f"{dns.get('score', 0):.2f} / 1.0")
        spf = dns.get("spf") or {}
        kv("SPF", f"[{spf.get('verdict', '?')}] {spf.get('value', '')}")
        dmarc = dns.get("dmarc") or {}
        kv("DMARC", f"[{dmarc.get('verdict', '?')}] {dmarc.get('value', '')}")
        kv("DKIM 힌트", dns.get("dkim_hint"))
        caa = dns.get("caa") or {}
        kv("CAA", f"[{caa.get('verdict', '?')}] {len(caa.get('records', []))}개 발견")
        for i, issue in enumerate(dns.get("issues") or [], 1):
            kv(f"이슈 {i}", issue)
        row += 1

    # TLS
    tls = evidence.get("tls") or {}
    if tls and not tls.get("error"):
        section_title("◆ TLS 인증서")
        kv("도메인", tls.get("domain"))
        kv("판정", tls.get("verdict"))
        kv("발급자", tls.get("issuer"))
        kv("주체", tls.get("subject"))
        kv("만료일", tls.get("not_after"))
        kv("남은 일수", tls.get("days_remaining"))
        kv("키 유형/길이", f"{tls.get('key_type', '?')} {tls.get('key_bits', '?')}bit")
        sans = tls.get("sans") or []
        kv("SAN", ", ".join(sans[:5]) + (f" 외 {len(sans)-5}개" if len(sans) > 5 else ""))
        for i, issue in enumerate(tls.get("issues") or [], 1):
            kv(f"이슈 {i}", issue)
        row += 1
    elif tls.get("error"):
        section_title("◆ TLS 인증서")
        kv("오류", tls.get("error"))
        row += 1

    # 공개 노출
    expo = evidence.get("exposure") or {}
    if expo and not expo.get("error"):
        section_title("◆ 공개 노출 (security.txt / robots.txt / .well-known/)")
        kv("기준 URL", expo.get("base_url"))
        kv("요약", expo.get("summary"))
        for c in (expo.get("checked") or []):
            status = c.get("status")
            path = c.get("path")
            if status == 200:
                kv(f"  · {path}", f"HTTP 200 ({c.get('size', 0)}B)")
            elif status:
                kv(f"  · {path}", f"HTTP {status}")
            else:
                kv(f"  · {path}", c.get("error", "?"))
        row += 1

    # GitHub repo
    gh = evidence.get("github") or {}
    if gh and not gh.get("error"):
        section_title("◆ GitHub repo 분석")
        kv("repo", f"{gh.get('owner')}/{gh.get('name')}")
        kv("default branch", gh.get("default_branch"))
        kv("Language", gh.get("language"))
        kv("License", gh.get("license"))
        kv("Stars", gh.get("stars"))
        kv("종합 점수", f"{gh.get('score', 0):.2f} / 1.0")
        for fname, exists in (gh.get("files") or {}).items():
            kv(f"  · {fname}", "있음" if exists else "없음")
        ci = gh.get("ci_workflows") or []
        kv("CI workflows", ", ".join(ci) if ci else "(없음)")
        for i, issue in enumerate(gh.get("issues") or [], 1):
            kv(f"이슈 {i}", issue)
    elif gh.get("error"):
        section_title("◆ GitHub repo 분석")
        kv("오류", gh.get("error"))


@router.get("/template/{session_id}")
def download_session_template(
    session_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """세션별 동적 양식 — 자동 폴백 항목까지 포함한 xlsx 생성·반환.

    SKT 가이드 §1·§5 권고: "자동수집이 안 되는 항목을 평가불가로 방치하지 말고
    수동 증적을 붙여 판정" 하기 위한 핵심 도구. T-Markov 처럼 Keycloak/Wazuh 환경이
    아닌 SaaS 형 평가에서는 이 동적 양식이 정적 양식보다 훨씬 많은 항목을 포함한다.
    """
    session = db.query(DiagnosisSession).filter(
        DiagnosisSession.session_id == session_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
    assert_session_access(current_user, session)

    try:
        data = _build_session_template_xlsx(session, db)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("[manual] session template build failed: %s", exc)
        raise HTTPException(status_code=500, detail="양식 생성에 실패했습니다.")

    filename = f"manual-checklist-session-{session_id}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/items/{session_id}")
def get_manual_items(
    session_id: int,
    excluded_tools: str = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """수동 진단 항목 + 미사용 도구 항목을 반환한다.
    excluded_tools: 쉼표 구분 도구명 (예: 'nmap,trivy') — 해당 도구 자동 항목도 수동으로 포함.

    추가: session.extra.profile_select 가 있으면 사용자가 쓰는 IdP/SIEM 외 자동 도구의
    체크리스트 항목을 자동으로 폴백 노출한다(예: idp_type='entra' → keycloak 항목 수동).
    """
    session = db.query(DiagnosisSession).filter(
        DiagnosisSession.session_id == session_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
    assert_session_access(current_user, session)

    excluded_set = {t.strip().lower() for t in excluded_tools.split(",") if t.strip()}
    # session.extra.profile_select 기반 자동 폴백 — _resolve_fallback_tools 와 동일 규칙.
    excluded_set |= _resolve_fallback_tools(session)
    excluded_list = sorted(excluded_set)
    if excluded_list:
        manual_items = db.query(Checklist).filter(
            or_(
                Checklist.diagnosis_type == "수동",
                Checklist.tool.in_(excluded_list),
            )
        ).all()
    else:
        manual_items = db.query(Checklist).filter(
            Checklist.diagnosis_type == "수동"
        ).all()

    submitted = {
        r.check_id
        for r in db.query(DiagnosisResult).filter(
            DiagnosisResult.session_id == session_id
        ).all()
    }

    return {
        "items": [
            {
                "check_id": item.check_id,
                "item_id": item.item_id,
                "pillar": item.pillar,
                "category": item.category,
                "item_name": item.item_name,
                "maturity": item.maturity,
                "criteria": item.criteria or "",
                "submitted": item.check_id in submitted,
            }
            for item in manual_items
        ],
        "total": len(manual_items),
        "submitted_count": len([i for i in manual_items if i.check_id in submitted]),
    }


# ─── P1-7: 수동 증적 파일 업로드 / 다운로드 ────────────────────────────────────


def _resolve_check_id(db: Session, check_id_raw) -> Optional[int]:
    """check_id 가 숫자(PK) 또는 item_id 문자열로 올 수 있도록 둘 다 지원."""
    if check_id_raw is None:
        return None
    if isinstance(check_id_raw, int):
        return check_id_raw
    s = str(check_id_raw).strip()
    if not s:
        return None
    if s.isdigit():
        return int(s)
    cl = db.query(Checklist).filter(Checklist.item_id == s).first()
    return cl.check_id if cl else None


@router.post("/upload-evidence")
async def upload_evidence(
    session_id: int = Form(...),
    check_id: str = Form(...),
    file: UploadFile = File(...),
    note: str = Form(""),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """진단 항목별 증적 파일 업로드 (PDF/이미지). 세션 권한 검증 후 로컬 디스크 저장."""
    session = db.query(DiagnosisSession).filter(
        DiagnosisSession.session_id == session_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
    assert_session_access(current_user, session)

    resolved_check_id = _resolve_check_id(db, check_id)
    if resolved_check_id is None:
        raise HTTPException(status_code=400, detail="check_id 가 유효하지 않습니다.")
    checklist = db.query(Checklist).filter(Checklist.check_id == resolved_check_id).first()
    if not checklist:
        raise HTTPException(status_code=404, detail="체크리스트 항목을 찾을 수 없습니다.")

    mime = (file.content_type or "").lower()
    if mime not in _ALLOWED_EVIDENCE_MIMES:
        raise HTTPException(
            status_code=400,
            detail=f"허용되지 않는 파일 형식입니다. (허용: {', '.join(sorted(_ALLOWED_EVIDENCE_MIMES))})",
        )

    max_bytes = MAX_EVIDENCE_SIZE_MB * 1024 * 1024
    content = await file.read()
    file_size = len(content)
    if file_size == 0:
        raise HTTPException(status_code=400, detail="빈 파일은 업로드할 수 없습니다.")
    if file_size > max_bytes:
        raise HTTPException(
            status_code=400,
            detail=f"파일 크기가 너무 큽니다. (최대 {MAX_EVIDENCE_SIZE_MB}MB)",
        )

    ext = _ALLOWED_EVIDENCE_MIMES[mime]
    target_dir = Path(EVIDENCE_STORAGE_DIR) / str(session_id) / str(resolved_check_id)
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        logger.error("[evidence] storage dir mkdir failed: %s", exc)
        raise HTTPException(status_code=500, detail="증적 저장소 초기화에 실패했습니다.")

    filename = f"{uuid.uuid4().hex}.{ext}"
    target_path = target_dir / filename
    try:
        with open(target_path, "wb") as fp:
            fp.write(content)
    except OSError as exc:
        logger.error("[evidence] write failed: %s", exc)
        raise HTTPException(status_code=500, detail="증적 파일 저장에 실패했습니다.")

    original_filename = (file.filename or "")[:255] or None
    note_str = (note or "").strip() or None

    existing = db.query(Evidence).filter(
        Evidence.session_id == session_id,
        Evidence.check_id == resolved_check_id,
    ).first()

    old_file_path: Optional[str] = None
    if existing:
        old_file_path = existing.file_path
        existing.source = "수동업로드"
        existing.observed = note_str
        existing.file_path = str(target_path)
        existing.mime_type = mime
        existing.file_size = file_size
        existing.original_filename = original_filename
        evidence = existing
    else:
        evidence = Evidence(
            session_id=session_id,
            check_id=resolved_check_id,
            source="수동업로드",
            observed=note_str,
            location="",
            reason="",
            impact=None,
            file_path=str(target_path),
            mime_type=mime,
            file_size=file_size,
            original_filename=original_filename,
        )
        db.add(evidence)

    db.commit()
    db.refresh(evidence)

    # 이전 파일이 있었다면 best-effort 삭제 (실패 무시 — 메타는 이미 갱신).
    if old_file_path and old_file_path != str(target_path):
        try:
            Path(old_file_path).unlink(missing_ok=True)
        except Exception as exc:
            logger.warning("[evidence] previous file unlink failed: %s", exc)

    return {
        "status":      "ok",
        "evidence_id": evidence.evidence_id,
        "file_path":   evidence.file_path,
        "file_size":   evidence.file_size,
        "mime_type":   evidence.mime_type,
        "original_filename": evidence.original_filename,
    }


@router.get("/evidence/{evidence_id}")
def download_evidence(
    evidence_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """증적 파일 다운로드. 세션 권한 검증."""
    evidence = db.query(Evidence).filter(Evidence.evidence_id == evidence_id).first()
    if not evidence:
        raise HTTPException(status_code=404, detail="증적을 찾을 수 없습니다.")
    if not evidence.file_path:
        raise HTTPException(status_code=404, detail="이 증적에는 첨부 파일이 없습니다.")

    session = db.query(DiagnosisSession).filter(
        DiagnosisSession.session_id == evidence.session_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
    assert_session_access(current_user, session)

    path = Path(evidence.file_path)
    if not path.is_file():
        logger.warning("[evidence] file missing on disk: %s", evidence.file_path)
        raise HTTPException(status_code=410, detail="증적 파일이 디스크에서 사라졌습니다.")

    return FileResponse(
        path=str(path),
        media_type=evidence.mime_type or "application/octet-stream",
        filename=evidence.original_filename or path.name,
    )
